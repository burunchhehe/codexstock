from __future__ import annotations

import tempfile
import unittest
import json
import time
import hashlib
import os
import threading
from unittest.mock import patch
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from codexstock_research_forge.models import ExperimentRecord, ExperimentStatus, StrategyDefinition
from codexstock_research_forge.service import ResearchForge
from codexstock_research_forge.gateway import call_research_tool
from codexstock_research_forge.adapters import AnalyticalStorageBacktestAdapter, LocalOhlcvBacktestAdapter
from codexstock_research_forge.microstructure import MicrostructureStore
from codexstock_research_forge.indicators import calculate, manifest, verify
from codexstock_research_forge.universe import KindOfficialUniverseProvider, KrxGlobalListingHistoryProvider, KrxOpenApiUniverseProvider, PointInTimeUniverseHistory, UniverseStore
from codexstock_research_forge.validation import validate_experiment_evidence
from codexstock_research_forge.collection import CollectionManager, KisReadOnlyProvider, MockMarketDataProvider
from codexstock_research_forge.execution import compare_execution_modes, execution_manifest, simulate_long_only
from codexstock_research_forge.storage import AnalyticalStorage
from codexstock_research_forge.jobs import AsyncJobManager
from codexstock_research_forge.analytics import optimized_walk_forward
from codexstock_research_forge.performance import run_performance_benchmark
from codexstock_research_forge.custom_indicators import CustomIndicatorRegistry
from codexstock_research_forge.multitimeframe import evaluate_multitimeframe
from codexstock_research_forge.soak import run_concurrency_soak
from codexstock_research_forge.microstructure_worker import MicrostructureWorker
from codexstock_research_forge.microstructure_archive import MicrostructureArchive
from codexstock_research_forge.hts_validation import HtsReferenceRegistry
from codexstock_research_forge.lifecycle import CONFIRMATION
from codexstock_research_forge.realtime import ORDERBOOK_COLUMNS, ORDERBOOK_TR_ID, TICK_COLUMNS, TICK_TR_ID, ReliableKisRealtimeCollector, parse_kis_frame, realtime_run_history
from codexstock_research_forge.corporate_actions import CorporateActionRegistry, OfficialCorporateActionEvidenceProvider, adjust_split_history
from codexstock_research_forge.regimes import analyze_market_regimes
from codexstock_research_forge.attribution import analyze_benchmark_attribution
from codexstock_research_forge.readiness import _complete_corporate_action_evidence, _complete_universe_evidence, _verified_performance, active_realtime_progress, qualified_full_session_runs
from codexstock_research_forge.hts_csv import import_csv as import_hts_csv, template as hts_csv_template
from codexstock_research_forge.instrument_contracts import (
    contract_manifest,
    normalize_instrument_dataset_contract,
    validate_instrument_snapshot,
)
from codexstock_research_forge.shards import ResearchShardCoordinator
from codexstock_research_forge.stability import EngineStabilityLedger


class ResearchForgeTests(unittest.TestCase):
    def test_engine_stability_ledger_hash_chain_and_regression_audit(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            ledger = EngineStabilityLedger(Path(directory) / "stability.jsonl")
            dashboard = {"generated_at": "2026-07-22T09:00:00+09:00", "engines": [{"engine_id": "vectorbt", "connected": True, "runtime_connected": True, "formal_connected": True, "adapter_ready": True, "current_usable": True, "operational_state": "normal", "root_cause_code": "none", "engine_commit": "abc", "runtime_mode": "on_demand", "live_order_allowed": False}]}
            self.assertTrue(ledger.record_dashboard(dashboard, 1)["recorded"])
            changed = json.loads(json.dumps(dashboard)); changed["generated_at"] = "2026-07-22T09:05:00+09:00"; changed["engines"][0]["engine_commit"] = "def"
            self.assertTrue(ledger.record_dashboard(changed, 1)["recorded"])
            audit = ledger.audit()
            self.assertEqual(audit["status"], "regression_detected")
            self.assertEqual(audit["regression_engine_ids"], ["vectorbt"])
            self.assertTrue(audit["audit_hash"].startswith("sha256:"))
            self.assertFalse(audit["live_order_allowed"])

    def test_engine_stability_does_not_claim_long_term_proof_from_two_snapshots(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            ledger = EngineStabilityLedger(Path(directory) / "stability.jsonl")
            engine = {"engine_id": "openbb", "connected": True, "runtime_connected": True, "formal_connected": True, "adapter_ready": True, "current_usable": True, "operational_state": "normal", "root_cause_code": "none", "engine_commit": "abc", "runtime_mode": "on_demand", "live_order_allowed": False}
            ledger.record_dashboard({"generated_at": "2026-07-22T09:00:00+09:00", "engines": [engine]}, 1)
            ledger.record_dashboard({"generated_at": "2026-07-22T09:05:00+09:00", "engines": [{**engine, "operational_state": "delayed"}]}, 1)
            audit = ledger.audit(window_days=30)
            self.assertEqual(audit["status"], "collecting_evidence")
            self.assertFalse(audit["enough_history"])
            self.assertGreater(audit["required_span_seconds"], audit["observed_span_seconds"])
    def test_distributed_shards_claim_exactly_once_and_hash_results(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            coordinator = ResearchShardCoordinator(Path(directory))
            created = coordinator.create_batch("walk_forward", [{"symbol": "005930"}, {"symbol": "000660"}])
            batch_id = created["manifest"]["batch_id"]
            first = coordinator.claim(batch_id, "worker-a", 30)
            self.assertTrue(first["claimed"])
            shard = first["shard"]
            coordinator.heartbeat(batch_id, shard["shard_id"], shard["worker_token"])
            completed = coordinator.finish(batch_id, shard["shard_id"], shard["worker_token"], result={"return_pct": 1.2})
            self.assertEqual(completed["shard"]["status"], "SUCCEEDED")
            second = coordinator.claim(batch_id, "worker-b", 30)
            self.assertTrue(second["claimed"])
            coordinator.finish(batch_id, second["shard"]["shard_id"], second["shard"]["worker_token"], error="transient", retryable=True)
            reclaimed = coordinator.claim(batch_id, "worker-c", 30)
            self.assertEqual(reclaimed["shard"]["shard_id"], second["shard"]["shard_id"])
            coordinator.finish(batch_id, reclaimed["shard"]["shard_id"], reclaimed["shard"]["worker_token"], result={"return_pct": 2.3})
            status = coordinator.status(batch_id)
            self.assertEqual(status["status_counts"], {"SUCCEEDED": 2})
            self.assertEqual(status["progress_percent"], 100.0)
            self.assertFalse(status["live_order_allowed"])
    def test_multi_market_instrument_contract_rejects_unit_and_currency_mixups(self) -> None:
        manifest_payload = contract_manifest()
        self.assertEqual(manifest_payload["contract_count"], 4)
        valid = validate_instrument_snapshot({"contract_id": "US_EQUITY", "symbol": "NVDA", "market": "US", "asset_class": "EQUITY", "currency": "USD", "quote_unit": "USD_PER_SHARE", "price": "202.81", "quantity": 2, "timestamp": "2026-07-22T10:00:00-04:00", "source": "fixture", "provider": "openbb"})
        self.assertTrue(valid["passed"])
        mixed = validate_instrument_snapshot({"contract_id": "KR_EQUITY", "symbol": "005930", "market": "KR", "asset_class": "EQUITY", "currency": "USD", "quote_unit": "USD_PER_SHARE", "price": 100, "quantity": 0.5, "timestamp": "2026-07-22T10:00:00", "source": "fixture"})
        self.assertFalse(mixed["passed"])
        self.assertIn("currency_contract_mismatch", mixed["errors"])
        self.assertIn("quote_unit_contract_mismatch", mixed["errors"])
        self.assertIn("fractional_quantity_not_allowed", mixed["errors"])
        self.assertIn("timestamp_timezone_missing", mixed["errors"])

    def test_experiment_enforces_instrument_contract_through_paper_gate(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory))
            record = forge.create_experiment(
                StrategyDefinition(
                    "contract-enforced",
                    "1",
                    {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8},
                ),
                {"dataset_id": "kr-fixture", "provider": "kis-readonly"},
                {"execution_mode": "REALISTIC"},
            )

            self.assertEqual("KR_EQUITY", record.data_snapshot["instrument_contract_id"])
            self.assertEqual("KRW", record.data_snapshot["currency"])
            self.assertEqual("KRW_PER_SHARE", record.data_snapshot["quote_unit"])
            self.assertTrue(record.data_snapshot["instrument_contract"]["passed"])

            readiness = forge.lifecycle_readiness(record.id)
            checks = {row["id"]: row["ok"] for row in readiness["checks"]}
            self.assertTrue(checks["instrument_contract_valid"])
            self.assertTrue(
                readiness["integrated_evidence"]["instrument_contract"]["passed"]
            )

            with self.assertRaisesRegex(ValueError, "currency_contract_mismatch"):
                forge.create_experiment(
                    StrategyDefinition(
                        "contract-rejected",
                        "1",
                        {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8},
                    ),
                    {
                        "dataset_id": "bad-fixture",
                        "currency": "USD",
                        "quote_unit": "USD_PER_SHARE",
                    },
                    {"execution_mode": "REALISTIC"},
                )

            mixed = normalize_instrument_dataset_contract(
                {"dataset_id": "mixed-fixture"},
                ["005930", "NVDA"],
            )
            self.assertFalse(mixed["passed"])
            self.assertIn("mixed_market_dataset_not_supported", mixed["errors"])
    def test_full_spec_evidence_requires_official_complete_provenance(self) -> None:
        evidence = {"official": True, "grade": "official_listing_interval_history", "coverage_start": "2000-01-01", "coverage_end": "2026-07-22", "includes_delisted": True, "complete_daily_history": True, "precoverage_listing_dates_clipped": 0}
        self.assertTrue(_complete_universe_evidence({"evidence": evidence}))
        self.assertFalse(_complete_universe_evidence({"evidence": {**evidence, "official": False}}))
        self.assertFalse(_complete_universe_evidence({"evidence": {**evidence, "coverage_start": ""}}))
        self.assertTrue(_complete_corporate_action_evidence({"complete_history": True, "source_documents_verified": True, "verified_source_document_count": 2}))
        self.assertFalse(_complete_corporate_action_evidence({"complete_history": True, "source_documents_verified": True, "verified_source_document_count": 0}))

    def test_doctor_enforces_research_only_policy(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory))
            self.assertTrue(forge.doctor()["ok"])
            self.assertFalse(forge.status()["policy"]["live_order_allowed"])

    def test_corporate_action_split_and_reverse_split_golden_adjustment(self) -> None:
        rows = [
            {"timestamp": "2024-01-01T00:00:00+09:00", "open": 100, "high": 100, "low": 100, "close": 100, "volume": 100},
            {"timestamp": "2024-01-02T00:00:00+09:00", "open": 100, "high": 100, "low": 100, "close": 100, "volume": 100},
            {"timestamp": "2024-01-03T00:00:00+09:00", "open": 50, "high": 50, "low": 50, "close": 50, "volume": 200},
            {"timestamp": "2024-01-04T00:00:00+09:00", "open": 250, "high": 250, "low": 250, "close": 250, "volume": 40},
        ]
        source_hash = "sha256:" + "a" * 64
        actions = [
            {"type": "SPLIT", "effective_date": "2024-01-03", "new_shares": 2, "old_shares": 1, "source_url": "https://kind.krx.co.kr/split", "source_hash": source_hash},
            {"type": "REVERSE_SPLIT", "effective_date": "2024-01-04", "new_shares": 1, "old_shares": 5, "source_url": "https://data.krx.co.kr/reverse", "source_hash": "sha256:" + "b" * 64},
        ]
        adjusted = adjust_split_history(rows, actions)
        self.assertEqual([row["close"] for row in adjusted["rows"]], [250.0] * 4)
        self.assertEqual([row["volume"] for row in adjusted["rows"]], [40.0] * 4)
        self.assertEqual(adjusted["ledger"][0]["price_factor_for_prior_rows"], 0.5)
        self.assertEqual(adjusted["ledger"][1]["price_factor_for_prior_rows"], 5.0)
        self.assertTrue(adjusted["adjusted_hash"].startswith("sha256:"))
        self.assertFalse(adjusted["live_order_allowed"])
        with self.assertRaisesRegex(ValueError, "already contain"):
            adjust_split_history(adjusted["rows"], actions)
        with self.assertRaisesRegex(ValueError, "official KRX"):
            adjust_split_history(rows, [{**actions[0], "source_url": "https://example.com/untrusted"}])

    def test_cash_dividend_and_rights_issue_golden_adjustment(self) -> None:
        rows = [
            {"timestamp": "2024-01-01T00:00:00+09:00", "open": 100, "high": 100, "low": 100, "close": 100, "volume": 100},
            {"timestamp": "2024-01-02T00:00:00+09:00", "open": 98, "high": 98, "low": 98, "close": 98, "volume": 100},
            {"timestamp": "2024-01-03T00:00:00+09:00", "open": 74, "high": 74, "low": 74, "close": 74, "volume": 200},
        ]
        actions = [
            {"type": "CASH_DIVIDEND", "effective_date": "2024-01-02", "cash_per_share": 2, "source_url": "https://kind.krx.co.kr/dividend", "source_hash": "sha256:" + "c" * 64},
            {"type": "RIGHTS_ISSUE", "effective_date": "2024-01-03", "new_shares": 1, "old_shares": 1, "subscription_price": 50, "source_url": "https://data.krx.co.kr/rights", "source_hash": "sha256:" + "d" * 64},
        ]
        adjusted = adjust_split_history(rows, actions)
        self.assertEqual([row["close"] for row in adjusted["rows"]], [74.0] * 3)
        self.assertEqual([row["volume"] for row in adjusted["rows"]], [200.0] * 3)
        self.assertEqual(adjusted["ledger"][0]["price_factor_for_prior_rows"], 0.98)
        self.assertEqual(adjusted["ledger"][0]["volume_factor_for_prior_rows"], 1.0)
        self.assertEqual(adjusted["ledger"][1]["price_factor_for_prior_rows"], round(74 / 98, 12))
        self.assertEqual(adjusted["ledger"][1]["volume_factor_for_prior_rows"], 2.0)
        with self.assertRaisesRegex(ValueError, "below the previous close"):
            adjust_split_history(rows, [{**actions[0], "cash_per_share": 100}])
        with self.assertRaisesRegex(ValueError, "subscription_price"):
            adjust_split_history(rows, [{**actions[1], "subscription_price": -1}])

    def test_market_regime_analysis_uses_only_trailing_prices(self) -> None:
        closes = [100.0] * 20 + [110.0] * 10 + [90.0] * 10 + [90.0] * 10
        rows = [{"date": (date(2024, 1, 1) + timedelta(days=index)).isoformat(), "close": close} for index, close in enumerate(closes)]
        equity = [1000 * (1.001 ** index) for index in range(len(rows))]
        result = analyze_market_regimes(rows, equity, lookback=10, threshold_pct=5)
        self.assertEqual(result["classified_bar_count"], 40)
        self.assertEqual(sum(value["bar_count"] for value in result["summary"].values()), 40)
        self.assertGreater(result["summary"]["BULL"]["bar_count"], 0)
        self.assertGreater(result["summary"]["BEAR"]["bar_count"], 0)
        self.assertGreater(result["summary"]["SIDEWAYS"]["bar_count"], 0)
        self.assertEqual(result["method"], "trailing_close_return_no_future_data")
        self.assertTrue(result["sequence_hash"].startswith("sha256:"))

    def test_benchmark_attribution_golden_identical_buy_and_hold(self) -> None:
        closes = [100.0, 102.0, 101.0, 104.0, 108.0, 107.0]
        rows = [{"date": (date(2024, 1, 1) + timedelta(days=index)).isoformat(), "close": close} for index, close in enumerate(closes)]
        result = analyze_benchmark_attribution(rows, [value * 10 for value in closes])
        self.assertAlmostEqual(result["strategy_return_pct"], result["benchmark_return_pct"], places=8)
        self.assertAlmostEqual(result["geometric_excess_return_pct"], 0.0, places=8)
        self.assertAlmostEqual(result["beta"], 1.0, places=8)
        self.assertAlmostEqual(result["correlation"], 1.0, places=8)
        self.assertAlmostEqual(result["annualized_tracking_error_pct"], 0.0, places=8)
        self.assertIsNone(result["information_ratio"])
        self.assertTrue(result["evidence_hash"].startswith("sha256:"))

    def test_corporate_action_registry_queries_adjusts_and_detects_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            registry = CorporateActionRegistry(Path(directory))
            actions = [{"type": "SPLIT", "effective_date": "2024-01-02", "new_shares": 2, "old_shares": 1, "source_url": "https://kind.krx.co.kr/split", "source_hash": "sha256:" + "e" * 64}]
            registered = registry.register("samsung-actions", "005930", actions, complete_history=False)
            self.assertEqual(registered["action_count"], 1)
            self.assertEqual(registry.query("samsung-actions", "2024-01-01", "2024-01-03")["count"], 1)
            self.assertEqual(registry.query("samsung-actions", "2025-01-01", "2025-01-03")["count"], 0)
            rows = [
                {"timestamp": "2024-01-01T00:00:00+09:00", "open": 100, "high": 100, "low": 100, "close": 100, "volume": 10},
                {"timestamp": "2024-01-02T00:00:00+09:00", "open": 50, "high": 50, "low": 50, "close": 50, "volume": 20},
            ]
            adjusted = registry.adjust("samsung-actions", rows)
            self.assertEqual([row["close"] for row in adjusted["rows"]], [50.0, 50.0])
            self.assertEqual(adjusted["registry_content_hash"], registered["content_hash"])
            self.assertFalse(adjusted["registry_complete_history"])
            path = Path(directory) / "samsung-actions.json"
            path.write_text(path.read_text(encoding="utf-8").replace('"new_shares": 2.0', '"new_shares": 3.0'), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "integrity"):
                registry.get("samsung-actions")
            self.assertFalse(registry.status()["ok"])

    def test_official_corporate_action_provider_hashes_exact_document_bytes(self) -> None:
        body = b"official-kind-document-v1"
        class Response:
            headers = {"Content-Type": "text/html;charset=UTF-8"}
            def __enter__(self): return self
            def __exit__(self, *args): return False
            def read(self, limit): return body
        opened = []
        def opener(request, timeout): opened.append((request.full_url, timeout)); return Response()
        provider = OfficialCorporateActionEvidenceProvider(opener)
        action = {"type": "SPLIT", "effective_date": "2024-01-02", "new_shares": 2, "old_shares": 1, "source_url": "https://kind.krx.co.kr/external/evidence.htm"}
        verified = provider.verify([action], timeout=3, max_bytes=1024, attempts=1)
        expected = "sha256:" + hashlib.sha256(body).hexdigest()
        self.assertEqual(verified["actions"][0]["source_hash"], expected)
        self.assertEqual(verified["documents"][0]["bytes"], len(body))
        self.assertEqual(opened, [(action["source_url"], 3)])
        self.assertTrue(verified["read_only"]); self.assertFalse(verified["order_allowed"])
        with self.assertRaisesRegex(ValueError, "does not match"):
            provider.verify([{**action, "source_hash": "sha256:" + "0" * 64}], max_bytes=1024, attempts=1)
        with self.assertRaisesRegex(ValueError, "KRX or KIND"):
            provider.verify([{**action, "source_url": "https://example.com/fake"}], max_bytes=1024, attempts=1)

    def test_corporate_action_complete_coverage_and_reconciliation_fail_closed(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            registry = CorporateActionRegistry(Path(directory))
            actions = [{"type": "SPLIT", "effective_date": "2024-01-02", "new_shares": 2, "old_shares": 1, "source_url": "https://kind.krx.co.kr/split", "source_hash": "sha256:" + "a" * 64}]
            with self.assertRaisesRegex(ValueError, "verified source documents"):
                registry.register("unverified-complete", "005930", actions, complete_history=True)
            registry.register("verified-complete", "005930", actions, complete_history=True, source_documents_verified=True, history_start="2000-01-01", history_end="2026-07-22")
            result = registry.reconcile(["005930", "000660"], "2000-01-01", "2026-07-22")
            self.assertFalse(result["passed"])
            self.assertEqual(result["missing_symbols"], ["000660"])
            self.assertTrue(result["evidence_hash"].startswith("sha256:"))
            self.assertFalse(result["order_allowed"])

    def test_readiness_separates_operational_from_research_and_full_spec_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); forge = ResearchForge.local(root)
            forge.storage().ingest([{
                "symbol": "005930", "timestamp": "2024-01-02T09:00:00+09:00", "open": 100,
                "high": 102, "low": 99, "close": 101, "volume": 1000, "source": "fixture",
            }], "1m")
            forge.microstructure().ingest([{
                "event_type": "tick", "symbol": "005930", "timestamp": "2024-01-02T09:00:00+09:00",
                "source": "fixture", "payload": {"price": 101, "volume": 1},
            }])
            forge.microstructure_archive().export_incremental()
            benchmarks = root / "benchmarks"; benchmarks.mkdir()
            run_performance_benchmark(forge.storage(), benchmarks, 1000)
            concurrency = {"schema_version": 1, "iterations": 2, "passed": True, "errors": [], "measurements": {name: {} for name in ("storage_write", "storage_read", "indicator_compute", "collection", "backtest")}}; canonical = json.dumps(concurrency, ensure_ascii=False, sort_keys=True, separators=(",", ":")); concurrency["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            (benchmarks / "concurrency_latest.json").write_text(json.dumps(concurrency), encoding="utf-8")

            failed_run_id = "realtime_run_" + "b" * 32; runs = root / "realtime" / "runs"; runs.mkdir(parents=True)
            failed_evidence = {"schema_version": 1, "status": "FAILED", "provider": "kis-readonly-websocket", "read_only": True, "order_allowed": False, "symbols": ["005930"], "stream_summaries": [], "run": {"run_id": failed_run_id, "started_at": "2024-01-01T00:00:00+00:00", "finished_at": "2024-01-01T00:00:01+00:00", "messages": 0, "data_messages": 0, "accepted_events": 0, "duplicate_events": 0, "reconnects": 0, "in_session_quality_ok": False}}
            canonical = json.dumps(failed_evidence, ensure_ascii=False, sort_keys=True, separators=(",", ":")); failed_evidence["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            (runs / f"{failed_run_id}.json").write_text(json.dumps(failed_evidence), encoding="utf-8")

            readiness = forge.readiness({"kis_configured": True, "websocket_available": True})
            self.assertTrue(readiness["engine_operational_ready"])
            self.assertFalse(readiness["research_evidence_ready"])
            self.assertFalse(readiness["full_spec_evidence_ready"])
            self.assertFalse(readiness["automatic_live_trading_allowed"])
            self.assertIn("strict_walk_forward_experiment", readiness["blockers"]["research"])
            self.assertIn("realtime_run_hashes", readiness["blockers"]["research"])
            self.assertEqual(readiness["evidence"]["realtime_runs"], 1)
            self.assertEqual(readiness["remaining_operational_blockers"], [])
            self.assertIn("strict_walk_forward_experiment", readiness["remaining_research_blockers"])
            self.assertIn("strict_walk_forward_experiment", readiness["remaining_full_spec_blockers"])
            self.assertIn("complete_historical_universe", readiness["remaining_full_spec_blockers"])
            self.assertIn("verified_corporate_action_history", readiness["remaining_full_spec_blockers"])
            self.assertEqual(readiness["evidence"]["corporate_action_dataset_count"], 0)
            self.assertEqual(readiness["evidence"]["verified_corporate_action_dataset_count"], 0)
            self.assertEqual(readiness["evidence"]["report_verification_errors"], [])
            record = ExperimentRecord(StrategyDefinition("diagnostic", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}), {"dataset_id": "fixture"}, {})
            forge.registry.save(record)
            with patch.object(forge, "verify_export", side_effect=RuntimeError("broken report fixture")):
                diagnosed = forge.readiness({"kis_configured": True, "websocket_available": True})
            self.assertEqual(diagnosed["evidence"]["report_verification_errors"][0]["experiment_id"], record.id)
            self.assertEqual(diagnosed["evidence"]["report_verification_errors"][0]["type"], "RuntimeError")
            first = ExperimentRecord(StrategyDefinition("first", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}), {"dataset_id": "fixture"}, {}, result={"row_count": 1})
            second = ExperimentRecord(StrategyDefinition("second", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}), {"dataset_id": "fixture"}, {}, result={"row_count": 1})
            forge.registry.save(first); forge.registry.save(second)
            with patch.object(forge, "verify_export", side_effect=lambda experiment_id: {"ok": experiment_id == first.id}):
                partial_reports = forge.readiness({"kis_configured": True, "websocket_available": True})
            report_check = next(row for row in partial_reports["checks"]["research"] if row["id"] == "verified_report_bundle")
            self.assertFalse(report_check["ok"])
            self.assertIn(second.id, report_check["detail"])

    def test_analytical_adapter_runs_ma_cross_and_pins_source_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); storage = AnalyticalStorage(root / "storage")
            rows = []
            for index in range(40):
                timestamp = (datetime(2024, 1, 1, tzinfo=timezone.utc) + timedelta(days=index)).isoformat()
                close = 100 + index
                rows.extend([
                    {"symbol": "005930", "timestamp": timestamp, "open": close, "high": close + 1, "low": close - 1, "close": close, "volume": 1000, "source": "official"},
                    {"symbol": "005930", "timestamp": timestamp.replace("00:00:00", "09:00:00"), "open": 10, "high": 11, "low": 9, "close": 10, "volume": 1, "source": "fixture"},
                ])
            storage.ingest(rows, "1d")
            adapter = AnalyticalStorageBacktestAdapter(root / "storage", root / "indicators")
            result = adapter.run(
                {"name": "source-pinned", "version": "1", "rules": {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}},
                {"dataset_id": "official-v1", "start_date": "2024-01-01", "end_date": "2024-02-20", "timeframe": "1d", "sources": ["official"]},
                {"execution_mode": "REALISTIC", "initial_cash": 1_000_000},
            )
            self.assertEqual(result["row_count"], 40)
            self.assertEqual(result["data_quality"]["sources"], ["official"])
            self.assertEqual(result["adapter"], "codexstock-analytical-multitimeframe-v1")
            excluded = adapter.run(
                {"name": "source-pinned", "version": "1", "rules": {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}},
                {"dataset_id": "official-v1", "start_date": "2024-01-01", "end_date": "2024-02-20", "timeframe": "1d", "sources": ["official"], "exclude_head_rows": 2, "exclude_tail_rows": 3},
                {"execution_mode": "REALISTIC", "initial_cash": 1_000_000},
            )
            self.assertEqual(excluded["row_count"], 35)
            self.assertEqual(excluded["data_quality"]["boundary_row_exclusions"], {"excluded_head_rows": 2, "excluded_tail_rows": 3})

    def test_official_interval_history_accepts_exact_requested_symbol_but_rejects_clipped_interval(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            store = UniverseStore(Path(directory))
            store.register("history", [
                {"symbol": "000660", "name": "exact", "market": "KOSPI", "security_type": "COMMON", "listing_date": "1996-12-26"},
                {"symbol": "999999", "name": "clipped", "market": "KOSPI", "security_type": "COMMON", "listing_date": "2016-07-06", "delisting_date": "2020-01-01"},
            ], "official", {"grade": "official_listing_interval_history", "coverage_start": "2016-07-06", "coverage_end": "2026-07-13", "precoverage_listing_dates_clipped": 1})
            self.assertTrue(store.validate_period("history", ["000660"], "2016-07-06", "2026-01-01")["passed"])
            clipped = store.validate_period("history", ["999999"], "2016-07-06", "2019-01-01")
            self.assertFalse(clipped["passed"])
            self.assertIn("evidence_grade", clipped["uncovered_symbols"])

    def test_experiment_is_persisted_and_runs_to_validation(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory))
            record = forge.create_experiment(
                StrategyDefinition("sma_cross", "1.0.0", {"entry": "SMA(5) > SMA(20)"}),
                {"dataset_id": "fixture-v1"},
                {"fill_model": "next_bar_open"},
            )
            completed = forge.run_backtest(record.id)
            self.assertEqual(completed.status, ExperimentStatus.VALIDATION)
            self.assertEqual(forge.registry.get(record.id).result["dataset_id"], "fixture-v1")
            self.assertEqual(forge.registry.get(record.id).backtest_adapter, "mock")

    def test_live_order_directive_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory))
            with self.assertRaisesRegex(ValueError, "live-order"):
                forge.create_experiment(
                    StrategyDefinition("unsafe", "1", {"live_order": True}),
                    {"dataset_id": "fixture-v1"},
                    {},
                )

    def test_typed_dsl_rejects_unknown_executable_surface(self) -> None:
        errors = StrategyDefinition(
            "unsafe_dsl", "1", {"type": "ma_cross", "symbol": "005930", "fast": 5, "slow": 20, "python": "open('x')"}
        ).validate()
        self.assertTrue(any("unknown ma_cross rule key: python" in error for error in errors))
        indicator_errors = StrategyDefinition(
            "unsafe_indicator", "1", {"type": "indicator_rules", "symbol": "005930", "entry": {"all": [{"left": {"python": "eval"}, "operator": ">", "right": 0}]}, "exit": {"all": [{"left": {"field": "close"}, "operator": "<", "right": 0}]}}
        ).validate()
        self.assertTrue(any("invalid indicator operand" in error for error in indicator_errors))
        for invalid, expected in ((-1, "between 0 and 10000"), (10001, "between 0 and 10000"), ("not-a-number", "must be an integer")):
            horizon_errors = StrategyDefinition(
                "invalid_horizon", "1", {"type": "ma_cross", "symbol": "005930", "fast": 5, "slow": 20, "label_horizon_rows": invalid}
            ).validate()
            self.assertTrue(any(expected in error for error in horizon_errors), (invalid, horizon_errors))

    def test_microstructure_checkpoint_duplicate_gap_and_restart(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = MicrostructureStore(root)
            event = {
                "event_type": "tick",
                "symbol": "005930",
                "timestamp": "2026-07-13T09:00:00+09:00",
                "source": "KIS",
                "payload": {"price": 70000, "quantity": 10},
            }
            first = store.ingest([event], gap_seconds=5)
            self.assertEqual(first["accepted"], 1)
            duplicate = store.ingest([event], gap_seconds=5)
            self.assertEqual(duplicate["duplicates"], 1)
            later = {**event, "timestamp": "2026-07-13T09:00:10+09:00", "payload": {"price": 70010, "quantity": 5}}
            gap = store.ingest([later], gap_seconds=5)
            self.assertEqual(len(gap["gaps"]), 1)
            restarted = MicrostructureStore(root)
            self.assertEqual(restarted.status()["total_accepted"], 2)
            self.assertEqual(restarted.quality()["gap_stream_count"], 1)
            with self.assertRaisesRegex(ValueError, "timestamp regression"):
                restarted.ingest([{**event, "timestamp": "2026-07-13T08:59:59+09:00"}], gap_seconds=5)
            self.assertEqual(restarted.status()["total_accepted"], 2)

    def test_microstructure_archive_exports_only_new_complete_lines_and_detects_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); store = MicrostructureStore(root / "microstructure")
            base = {"event_type": "tick", "symbol": "005930", "source": "fixture", "payload": {"price": 100, "volume": 1}}
            store.ingest([{**base, "timestamp": "2024-01-02T09:00:00+09:00"}, {**base, "timestamp": "2024-01-02T09:00:01+09:00", "payload": {"price": 101, "volume": 2}}])
            archive = MicrostructureArchive(root / "microstructure")
            first = archive.export_incremental()
            self.assertEqual(first["exported_rows"], 2)
            self.assertEqual(archive.export_incremental()["exported_chunk_count"], 0)
            store.ingest([{**base, "timestamp": "2024-01-02T09:00:02+09:00", "payload": {"price": 102, "volume": 3}}])
            second = archive.export_incremental()
            self.assertEqual(second["exported_rows"], 1)
            self.assertEqual(second["status"]["row_count"], 3)
            self.assertTrue(archive.verify()["verified"])
            queried = archive.query(["005930"], "2024-01-02T00:00:00+09:00", "2024-01-03T00:00:00+09:00", ["tick"], 2)
            self.assertEqual(queried["count"], 2)
            self.assertEqual(queried["matched"], 3)
            self.assertTrue(queried["truncated"])
            self.assertEqual(queried["events"][0]["payload"]["price"], 100)
            parquet = next((root / "microstructure" / "archive" / "parquet").rglob("*.parquet"))
            with parquet.open("ab") as handle: handle.write(b"tamper")
            self.assertFalse(archive.verify()["verified"])

    def test_microstructure_worker_is_read_only_persistent_and_deduplicates_polls(self) -> None:
        class Provider:
            name = "fixture-poll"; read_only = True
            def status(self): return {"configured": True, "order_allowed": False, "capabilities": {"tick": True, "orderbook": True, "program_flow": False}}
            def poll(self, symbol):
                return {"ok": True, "events": [
                    {"event_type": "tick", "symbol": symbol, "timestamp": "2024-01-02T09:00:00+09:00", "source": "fixture_poll", "payload": {"price": 100, "volume": 1, "strength": 110}},
                    {"event_type": "orderbook", "symbol": symbol, "timestamp": "2024-01-02T09:00:00+09:00", "source": "fixture_poll", "payload": {"levels": [{"ask_price": 101, "bid_price": 99}]}}], "availability": {"tick": True, "orderbook": True, "program_flow": False}, "errors": {"program_flow": "unavailable"}}
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); store = MicrostructureStore(root / "store"); worker = MicrostructureWorker(root / "workers", store)
            result = worker.start(Provider(), ["005930"], interval_seconds=0.05, max_cycles=2)
            self.assertTrue(result["ok"])
            self.assertEqual(result["worker"]["accepted_events"], 2)
            self.assertEqual(result["worker"]["duplicate_events"], 2)
            self.assertFalse(result["worker"]["availability"]["005930"]["program_flow"])
            recovered = MicrostructureWorker(root / "workers", MicrostructureStore(root / "store")).get(result["worker"]["worker_id"])
            self.assertEqual(recovered["status"], "COMPLETED")
            class Unsafe(Provider): read_only = False
            with self.assertRaisesRegex(ValueError, "read-only"):
                worker.start(Unsafe(), ["005930"])

    def test_kis_websocket_parser_reconnects_restores_subscriptions_and_deduplicates(self) -> None:
        tick = {column: "0" for column in TICK_COLUMNS}
        tick.update({"symbol": "005930", "time": "090001", "business_date": "20240701", "price": "80000", "volume": "3", "accumulated_volume": "1003", "ask": "80100", "bid": "80000", "strength": "121.5"})
        tick_frame = f"0|{TICK_TR_ID}|1|" + "^".join(tick[column] for column in TICK_COLUMNS)
        orderbook = {column: "0" for column in ORDERBOOK_COLUMNS}
        orderbook.update({"symbol": "005930", "time": "090002", "ask_1": "80100", "ask_quantity_1": "10", "bid_1": "80000", "bid_quantity_1": "12", "total_ask_quantity": "100", "total_bid_quantity": "120", "kmid_price": "80050", "kmid_total_quantity": "22", "kmid_code": "A"})
        orderbook_frame = f"0|{ORDERBOOK_TR_ID}|1|" + "^".join(orderbook[column] for column in ORDERBOOK_COLUMNS)
        parsed = parse_kis_frame(tick_frame)
        self.assertEqual(parsed["events"][0]["payload"]["price"], 80000)
        self.assertEqual(parsed["events"][0]["timestamp"], "2024-07-01T09:00:01+09:00")

        class FakeTransport:
            def __init__(self, frames): self.frames, self.sent, self.closed = list(frames), [], False
            def send(self, value): self.sent.append(value)
            def recv(self, timeout):
                value = self.frames.pop(0)
                if isinstance(value, Exception): raise value
                return value
            def close(self): self.closed = True

        first = FakeTransport([ConnectionError("injected disconnect")])
        heartbeat = json.dumps({"header": {"tr_id": "PINGPONG"}})
        second = FakeTransport([heartbeat, tick_frame, tick_frame, orderbook_frame])
        transports = [first, second]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); store = MicrostructureStore(root / "store")
            collector = ReliableKisRealtimeCollector(root / "collector", store, lambda: transports.pop(0), "approval-fixture")
            result = collector.run(["005930"], max_messages=4, max_reconnects=2, heartbeat_timeout=10)
            self.assertTrue(result["ok"])
            self.assertEqual(result["collector"]["reconnect_count"], 1)
            self.assertEqual(result["collector"]["subscription_restore_count"], 4)
            self.assertEqual(result["collector"]["heartbeat_count"], 1)
            self.assertEqual(result["collector"]["accepted_events"], 2)
            self.assertEqual(result["collector"]["duplicate_events"], 1)
            self.assertTrue(result["run_quality"]["in_session_quality_ok"])
            self.assertEqual(result["run_quality"]["duplicate_events"], 1)
            self.assertTrue(result["run_quality"]["reconnect_recovery_ok"])
            self.assertEqual(result["run_quality"]["recovered_errors"], 1)
            self.assertIsNone(result["collector"]["last_error"])
            self.assertEqual(result["collector"]["last_recovered_error"]["type"], "ConnectionError")
            self.assertEqual(result["run_quality"]["connections"], 2)
            self.assertEqual(result["run_quality"]["subscription_restores"], 4)
            self.assertEqual(result["run_quality"]["completion_reason"], "message_limit_reached")
            self.assertEqual(parse_kis_frame(orderbook_frame)["events"][0]["payload"]["kmid_price"], 80050)
            self.assertEqual(len(first.sent), 2)
            self.assertEqual(len(second.sent), 3)
            self.assertTrue((root / "collector" / "checkpoint.json").is_file())
            evidence = Path(result["collector"]["last_run_evidence"]["path"])
            self.assertTrue(evidence.is_file())
            verified_history = realtime_run_history(root / "collector")
            self.assertTrue(verified_history["verified"])
            self.assertEqual(verified_history["completed_data_run_count"], 1)
            self.assertEqual(verified_history["status_counts"], {"COMPLETED": 1})
            self.assertTrue(verified_history["runs"][0]["evidence_verified"])
            self.assertEqual(verified_history["runs"][0]["quality_scope"]["run_quality_fields"], "current_run_only")
            self.assertEqual(verified_history["runs"][0]["quality_scope"]["stream_summaries"], "cumulative_store_state_at_run_finish")
            retransmitted = json.loads(evidence.read_text(encoding="utf-8")); retransmitted.pop("evidence_hash")
            retransmitted["stream_summaries"][0]["duplicate_count"] = retransmitted["stream_summaries"][0]["count"] + 5
            canonical = json.dumps(retransmitted, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            retransmitted["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence.write_text(json.dumps(retransmitted), encoding="utf-8")
            self.assertTrue(realtime_run_history(root / "collector")["verified"])
            false_elapsed = json.loads(json.dumps(retransmitted)); false_elapsed.pop("evidence_hash"); false_elapsed["run"]["elapsed_seconds"] = 14_400
            canonical = json.dumps(false_elapsed, ensure_ascii=False, sort_keys=True, separators=(",", ":")); false_elapsed["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence.write_text(json.dumps(false_elapsed), encoding="utf-8")
            self.assertFalse(realtime_run_history(root / "collector")["verified"])
            string_counter = json.loads(json.dumps(retransmitted)); string_counter.pop("evidence_hash"); string_counter["run"]["messages"] = str(string_counter["run"]["messages"])
            canonical = json.dumps(string_counter, ensure_ascii=False, sort_keys=True, separators=(",", ":")); string_counter["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence.write_text(json.dumps(string_counter), encoding="utf-8")
            self.assertFalse(realtime_run_history(root / "collector")["verified"])
            wrong_stream = json.loads(json.dumps(retransmitted)); wrong_stream.pop("evidence_hash")
            wrong_stream["stream_summaries"][0]["stream_id"] = "kis_readonly_websocket:tick:999999"
            canonical = json.dumps(wrong_stream, ensure_ascii=False, sort_keys=True, separators=(",", ":")); wrong_stream["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence.write_text(json.dumps(wrong_stream), encoding="utf-8")
            self.assertFalse(realtime_run_history(root / "collector")["verified"])
            naive_time = json.loads(json.dumps(retransmitted)); naive_time.pop("evidence_hash"); naive_time["run"]["started_at"] = naive_time["run"]["started_at"].replace("+00:00", "")
            canonical = json.dumps(naive_time, ensure_ascii=False, sort_keys=True, separators=(",", ":")); naive_time["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence.write_text(json.dumps(naive_time), encoding="utf-8")
            self.assertFalse(realtime_run_history(root / "collector")["verified"])
            evidence.write_text(json.dumps(retransmitted), encoding="utf-8")
            structurally_tampered = json.loads(evidence.read_text(encoding="utf-8"))
            structurally_tampered["stream_summaries"][0]["count"] = -1
            structurally_tampered.pop("evidence_hash")
            canonical = json.dumps(structurally_tampered, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            structurally_tampered["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence.write_text(json.dumps(structurally_tampered), encoding="utf-8")
            self.assertFalse(realtime_run_history(root / "collector")["verified"])
            evidence.write_text(evidence.read_text(encoding="utf-8").replace("approval", "altered"), encoding="utf-8")
            # The fixture contains no credential text; mutate a hashed field explicitly.
            payload = json.loads(evidence.read_text(encoding="utf-8")); payload["status"] = "ALTERED"; evidence.write_text(json.dumps(payload), encoding="utf-8")
            self.assertFalse(realtime_run_history(root / "collector")["verified"])

    def test_kis_websocket_duration_bound_finishes_without_message_limit(self) -> None:
        class HeartbeatTransport:
            def send(self, value): pass
            def recv(self, timeout):
                time.sleep(min(0.01, timeout))
                return json.dumps({"header": {"tr_id": "PINGPONG"}})
            def close(self): pass
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); collector = ReliableKisRealtimeCollector(root / "collector", MicrostructureStore(root / "store"), HeartbeatTransport, "approval-fixture")
            result = collector.run(["005930"], max_messages=0, duration_seconds=0.04, heartbeat_timeout=1)
            self.assertTrue(result["ok"])
            self.assertGreater(result["run_quality"]["messages"], 0)
            self.assertEqual(result["run_quality"]["requested_duration_seconds"], 0.04)
            self.assertEqual(result["collector"]["requested_duration_seconds"], 0.04)
            self.assertEqual(result["collector"]["run_message_count"], result["run_quality"]["messages"])
            self.assertGreaterEqual(result["run_quality"]["elapsed_seconds"], 0.035)
            self.assertTrue(result["run_quality"]["duration_completed"])
            self.assertEqual(result["run_quality"]["completion_reason"], "duration_reached")
            self.assertTrue(result["run_quality"]["continuous_quality_ok"])

    def test_realtime_history_rejects_self_hashed_but_structurally_empty_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); runs = root / "runs"; runs.mkdir()
            run_id = "realtime_run_" + "a" * 32
            payload = {
                "schema_version": 1, "status": "COMPLETED", "provider": "kis-readonly-websocket",
                "read_only": True, "order_allowed": False, "symbols": ["005930"],
                "stream_summaries": [], "run": {"run_id": run_id},
            }
            canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            payload["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            (runs / f"{run_id}.json").write_text(json.dumps(payload), encoding="utf-8")
            history = realtime_run_history(root)
            self.assertFalse(history["verified"])
            self.assertEqual(history["run_count"], 0)
            self.assertEqual(history["invalid_files"], [f"{run_id}.json"])

    def test_kis_websocket_resume_preserves_lineage_and_only_runs_remaining_duration(self) -> None:
        class HeartbeatTransport:
            def send(self, value): pass
            def recv(self, timeout):
                time.sleep(min(0.005, timeout))
                return json.dumps({"header": {"tr_id": "PINGPONG"}})
            def close(self): pass
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); collector_root = root / "collector"; collector_root.mkdir()
            previous_id = "realtime_run_" + "a" * 32
            (collector_root / "checkpoint.json").write_text(json.dumps({
                "status": "FAILED", "symbols": ["005930"], "run_id": previous_id,
                "session_chain_id": previous_id, "requested_duration_seconds": 0.05,
                "last_run": {"run_id": previous_id, "session_chain_id": previous_id, "requested_duration_seconds": 0.05, "elapsed_seconds": 0.02},
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }), encoding="utf-8")
            collector = ReliableKisRealtimeCollector(collector_root, MicrostructureStore(root / "store"), HeartbeatTransport, "approval-fixture")
            result = collector.resume_interrupted(heartbeat_timeout=1)
            self.assertTrue(result["ok"])
            self.assertEqual(result["resume"]["resumed_from_run_id"], previous_id)
            self.assertEqual(result["run_quality"]["session_chain_id"], previous_id)
            self.assertEqual(result["run_quality"]["resumed_from_run_id"], previous_id)
            self.assertAlmostEqual(result["run_quality"]["requested_duration_seconds"], 0.03, places=3)
            history = realtime_run_history(collector_root)
            self.assertTrue(history["verified"])
            self.assertEqual(history["run_count"], 1)

    def test_full_session_gate_requires_actual_elapsed_data_and_reconnect_recovery(self) -> None:
        run = {"requested_duration_seconds": 14_400, "elapsed_seconds": 14_400, "duration_completed": True, "completion_reason": "duration_reached", "reconnect_recovery_ok": True, "data_messages": 100, "in_session_quality_ok": True}
        row = {"status": "COMPLETED", "provider": "kis-readonly-websocket", "read_only": True, "order_allowed": False, "run": run}
        self.assertEqual(qualified_full_session_runs({"runs": [row]}), [row])
        for key, bad in (("elapsed_seconds", 100), ("duration_completed", False), ("reconnect_recovery_ok", False), ("data_messages", 0), ("in_session_quality_ok", False)):
            candidate = {**row, "run": {**run, key: bad}}
            self.assertEqual(qualified_full_session_runs({"runs": [candidate]}), [], key)
        wrong_reason = {**row, "run": {**run, "completion_reason": "message_limit_reached"}}
        self.assertEqual(qualified_full_session_runs({"runs": [wrong_reason]}), [])
        for key, bad in (("provider", "unknown"), ("read_only", False), ("order_allowed", True)):
            candidate = {**row, key: bad}
            self.assertEqual(qualified_full_session_runs({"runs": [candidate]}), [], key)
        legacy = {"status": "COMPLETED", "provider": "kis-readonly-websocket", "read_only": True, "order_allowed": False, "run": {"requested_duration_seconds": 14_400, "started_at": "2024-01-01T00:00:00+00:00", "finished_at": "2024-01-01T04:00:01+00:00", "reconnects": 0, "data_messages": 100, "in_session_quality_ok": True}}
        self.assertEqual(qualified_full_session_runs({"runs": [legacy]}), [legacy])
        legacy["run"]["reconnects"] = 1
        self.assertEqual(qualified_full_session_runs({"runs": [legacy]}), [])

    def test_readiness_reports_fresh_active_realtime_progress(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); root.mkdir(exist_ok=True)
            now = datetime.now(timezone.utc)
            (root / "checkpoint.json").write_text(json.dumps({
                "status": "RUNNING", "run_id": "run-active", "run_started_at": (now - timedelta(seconds=30)).isoformat(), "requested_duration_seconds": 120,
                "updated_at": now.isoformat(), "run_message_count": 10, "run_data_messages": 8, "run_accepted_events": 21, "run_duplicate_events": 1, "run_reconnect_count": 2,
                "data_messages": 12, "accepted_events": 34, "duplicate_events": 3,
                "last_error": None, "read_only": True, "order_allowed": False,
            }), encoding="utf-8")
            progress = active_realtime_progress(root)
            self.assertTrue(progress["fresh"])
            self.assertGreaterEqual(progress["elapsed_seconds"], 29)
            self.assertEqual(progress["run_data_messages"], 8)
            self.assertEqual(progress["run_accepted_events"], 21)
            self.assertEqual(progress["cumulative_data_messages"], 12)
            self.assertEqual(progress["counter_scope"], "current_run")
            self.assertGreaterEqual(progress["duration_progress_percent"], 24)
            self.assertLess(progress["duration_progress_percent"], 26)
            self.assertTrue(progress["read_only"])
            self.assertFalse(progress["order_allowed"])

    def test_kis_websocket_singleton_rejects_live_collector_and_reclaims_dead_lease(self) -> None:
        class HeartbeatTransport:
            def send(self, value): pass
            def recv(self, timeout): time.sleep(min(0.01, timeout)); return json.dumps({"header": {"tr_id": "PINGPONG"}})
            def close(self): pass
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); collector_root = root / "collector"; collector_root.mkdir()
            checkpoint = collector_root / "checkpoint.json"
            checkpoint.write_text(json.dumps({"status": "RUNNING", "updated_at": datetime.now(timezone.utc).isoformat()}), encoding="utf-8")
            collector = ReliableKisRealtimeCollector(collector_root, MicrostructureStore(root / "store"), HeartbeatTransport, "approval-fixture")
            with self.assertRaisesRegex(RuntimeError, "fresh RUNNING"):
                collector.run(["005930"], duration_seconds=0.02)
            checkpoint.write_text(json.dumps({"status": "RUNNING", "updated_at": "2020-01-01T00:00:00+00:00"}), encoding="utf-8")
            lease = collector_root / "collector.lease"; lease.mkdir(); (lease / "owner.json").write_text(json.dumps({"pid": 99999999}), encoding="utf-8")
            result = collector.run(["005930"], duration_seconds=0.02)
            self.assertTrue(result["ok"])
            self.assertFalse(lease.exists())

    def test_indicator_reference_values_and_profile_verification(self) -> None:
        rows = [
            {"open": value, "high": value + 1, "low": value - 1, "close": value, "volume": 100}
            for value in (1.0, 2.0, 3.0, 4.0, 5.0)
        ]
        sma = calculate("SMA", rows, {"period": 3})
        ema = calculate("EMA", rows, {"period": 3})
        rsi = calculate("RSI", rows, {"period": 3})
        self.assertEqual(sma["outputs"]["value"][-1], 4.0)
        self.assertEqual(ema["outputs"]["value"][-1], 4.0)
        self.assertEqual(rsi["outputs"]["value"][-1], 100.0)
        checked = verify("SMA", rows, {"period": 3}, {"value": 4.0}, "LS_HTS", tolerance=0)
        self.assertTrue(checked["passed"])
        profiles = {row["name"]: row for row in manifest()["profiles"]}
        self.assertFalse(profiles["LS_HTS"]["compatibility_verified"])

    def test_extended_indicator_golden_properties(self) -> None:
        rows = [{"open": float(value), "high": float(value + 1), "low": float(value - 1), "close": float(value), "volume": 100.0} for value in range(1, 61)]
        self.assertAlmostEqual(calculate("WMA", rows, {"period": 3})["outputs"]["value"][-1], 59.3333333333, places=8)
        self.assertAlmostEqual(calculate("ROC", rows, {"period": 10})["outputs"]["value"][-1], 20.0, places=8)
        self.assertEqual(calculate("OBV", rows, {})["outputs"]["value"][-1], 5900.0)
        self.assertAlmostEqual(calculate("VWAP", rows, {})["outputs"]["value"][-1], 30.5, places=8)
        stochastic = calculate("STOCHASTIC", rows, {"period": 14, "smooth": 3})["outputs"]
        williams = calculate("WILLIAMS_R", rows, {"period": 14})["outputs"]["value"]
        self.assertAlmostEqual(williams[-1], stochastic["k"][-1] - 100.0, places=8)
        self.assertEqual(calculate("MFI", rows, {"period": 14})["outputs"]["value"][-1], 100.0)
        self.assertIsNotNone(calculate("CCI", rows, {"period": 20})["outputs"]["value"][-1])
        self.assertIsNotNone(calculate("MACD", rows, {"fast": 12, "slow": 26, "signal": 9})["outputs"]["histogram"][-1])
        self.assertAlmostEqual(calculate("ATR", rows, {"period": 14})["outputs"]["value"][-1], 2.0, places=8)
        self.assertAlmostEqual(calculate("ADX", rows, {"period": 14})["outputs"]["adx"][-1], 100.0, places=8)

    def test_wilder_rsi_matches_fixed_published_golden_vector(self) -> None:
        fixture = json.loads((Path(__file__).parent / "fixtures" / "wilder_rsi_golden.json").read_text(encoding="utf-8"))
        rows = [{"open": value, "high": value, "low": value, "close": value, "volume": 1} for value in fixture["closes"]]
        actual = calculate("RSI", rows, {"period": fixture["period"]})["outputs"]["value"][fixture["period"]:]
        self.assertEqual(len(actual), len(fixture["expected_from_index_14"]))
        for value, expected in zip(actual, fixture["expected_from_index_14"]):
            self.assertAlmostEqual(value, expected, delta=fixture["absolute_tolerance"])

    def test_hts_reference_evidence_is_immutable_and_promotes_only_matching_indicator(self) -> None:
        fixture = json.loads((Path(__file__).parent / "fixtures" / "wilder_rsi_golden.json").read_text(encoding="utf-8"))
        first = date(2024, 1, 1)
        rows = [{"date": (first + timedelta(days=index)).isoformat(), "open": value, "high": value, "low": value, "close": value, "volume": 100} for index, value in enumerate(fixture["closes"])]
        references = [{"timestamp": rows[index + fixture["period"]]["date"], "outputs": {"value": expected}} for index, expected in enumerate(fixture["expected_from_index_14"])]
        package = {"export_id": "ls-rsi-fixture-1", "profile": "LS_HTS", "indicator": "RSI", "parameters": {"period": 14}, "symbol": "005930", "timeframe": "1d", "exported_at": "2024-02-10T15:40:00+09:00", "source_file_name": "ls_rsi_export.csv", "source_file_hash": "sha256:" + "1" * 64, "market_rows": rows, "reference_points": references, "absolute_tolerance": fixture["absolute_tolerance"]}
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory)); registry = forge.hts_references()
            passed = registry.register(package)
            self.assertTrue(passed["passed"])
            status = registry.status("LS_HTS")
            self.assertEqual(status["profiles"]["LS_HTS"]["verified_indicators"], ["RSI"])
            self.assertFalse(status["profiles"]["LS_HTS"]["fully_verified"])
            evidence_path = Path(directory) / "hts_references" / "ls-rsi-fixture-1.json"
            tampered = json.loads(evidence_path.read_text(encoding="utf-8")); tampered["verification"]["passed"] = False
            evidence_path.write_text(json.dumps(tampered), encoding="utf-8")
            invalid = registry.status("LS_HTS")
            self.assertFalse(invalid["ok"])
            self.assertEqual(invalid["package_count"], 0)
            self.assertEqual(len(invalid["invalid"]), 1)
            doctor_check = next(row for row in forge.doctor()["checks"] if row["id"] == "hts_reference_registry")
            self.assertFalse(doctor_check["ok"])
            with self.assertRaisesRegex(ValueError, "integrity"):
                registry.register(package)

    def test_hts_csv_template_import_and_registry_verification(self) -> None:
        generated = hts_csv_template("LS_HTS", "SMA")
        self.assertEqual(generated["columns"], ["timestamp", "open", "high", "low", "close", "volume", "hts_value"])
        lines = [generated["csv_header"].strip()]
        for index in range(25):
            close = 100 + index; expected = "" if index < 2 else str(close - 1)
            lines.append(f"2024-01-{index + 1:02d}T00:00:00+09:00,{close},{close + 1},{close - 1},{close},1000,{expected}")
        csv_text = "\n".join(lines) + "\n"
        metadata = {"export_id": "ls-sma-csv-1", "profile": "LS_HTS", "indicator": "SMA", "parameters": {"period": 3}, "symbol": "005930", "timeframe": "1d", "exported_at": "2024-02-01T15:30:00+09:00", "source_file_name": "ls_sma.csv", "absolute_tolerance": 0}
        imported = import_hts_csv(csv_text, metadata)
        self.assertEqual(imported["row_count"], 25)
        self.assertEqual(imported["reference_row_count"], 23)
        self.assertEqual(imported["source_file_hash"], "sha256:" + hashlib.sha256(csv_text.encode()).hexdigest())
        with tempfile.TemporaryDirectory() as directory:
            evidence = HtsReferenceRegistry(Path(directory)).register(imported["package"])
            self.assertTrue(evidence["passed"])
        with self.assertRaisesRegex(ValueError, "header"):
            import_hts_csv(csv_text.replace("hts_value", "value", 1), metadata)
        duplicate = csv_text.replace("2024-01-02T00:00:00+09:00", "2024-01-01T00:00:00+09:00", 1)
        with self.assertRaisesRegex(ValueError, "unique, increasing"):
            import_hts_csv(duplicate, metadata)
            failed_refs = json.loads(json.dumps(references)); failed_refs[-1]["outputs"]["value"] += 1
            failed = registry.register({**package, "export_id": "ls-rsi-fixture-bad", "reference_points": failed_refs, "source_file_hash": "sha256:" + "2" * 64})
            self.assertFalse(failed["passed"])
            self.assertEqual(failed["verification"]["mismatch_count"], 1)
            with self.assertRaisesRegex(ValueError, "immutable"):
                registry.register({**package, "notes": "changed"})

    def test_custom_indicator_requires_golden_determinism_and_prefix_stability(self) -> None:
        rows = [{"open": value, "high": value + 1, "low": value - 1, "close": value, "volume": 100} for value in range(1, 8)]
        definition = {
            "name": "CLOSE_MINUS_SMA", "version": "1.0", "description": "distance from SMA(3)",
            "outputs": {"value": {"op": "SUBTRACT", "args": [{"field": "close"}, {"indicator": "SMA", "parameters": {"period": 3}, "output": "value"}]}},
            "golden_tests": [{"rows": rows, "expected_last": {"value": 1.0}, "tolerance": 0}],
        }
        with tempfile.TemporaryDirectory() as directory:
            registry = CustomIndicatorRegistry(Path(directory))
            registered = registry.register(definition)
            self.assertTrue(registered["verification"]["prefix_stability_checked"])
            calculation = registry.calculate("CLOSE_MINUS_SMA", "1.0", rows)
            self.assertEqual(calculation["outputs"]["value"][-1], 1.0)
            self.assertTrue(calculation["result_hash"].startswith("sha256:"))
            with self.assertRaisesRegex(ValueError, "immutable"):
                registry.register({**definition, "description": "changed"})
            bad = {**definition, "name": "BAD_GOLDEN", "golden_tests": [{"rows": rows, "expected_last": {"value": 999}}]}
            with self.assertRaisesRegex(ValueError, "verification failed"):
                registry.register(bad)

    def test_multitimeframe_alignment_waits_for_completed_higher_bar(self) -> None:
        daily = [
            {"timestamp": "2024-01-01T00:00:00+09:00", "open": 10, "high": 12, "low": 9, "close": 10, "volume": 1000},
            {"timestamp": "2024-01-02T00:00:00+09:00", "open": 100, "high": 102, "low": 99, "close": 100, "volume": 1000},
        ]
        minute = []
        for timestamp in ["2024-01-01T09:00:00+09:00", "2024-01-01T09:01:00+09:00", "2024-01-02T09:00:00+09:00", "2024-01-02T09:01:00+09:00", "2024-01-02T09:02:00+09:00"]:
            minute.append({"timestamp": timestamp, "open": 20, "high": 21, "low": 19, "close": 20, "volume": 1000})
        rules = {
            "type": "multi_timeframe_indicator_rules", "symbol": "005930", "execution_context": "intraday",
            "contexts": {"daily": {"timeframe": "1d"}, "intraday": {"timeframe": "1m"}},
            "entry": {"all": [{"left": {"field": "close", "timeframe": "daily"}, "operator": "<", "right": 50}]},
            "exit": {"all": [{"left": {"field": "close", "timeframe": "daily"}, "operator": ">", "right": 50}]},
        }
        evaluated = evaluate_multitimeframe(rules, {"daily": daily, "intraday": minute})
        self.assertEqual(evaluated["entry_signals"], [False, False, True, True, True])
        self.assertEqual(evaluated["exit_signals"], [False] * 5)
        self.assertEqual(evaluated["future_information_violations"], 0)
        self.assertEqual(evaluated["alignment"]["daily"]["unavailable_prefix_rows"], 2)

    def test_strict_walk_forward_never_uses_test_window_for_selection(self) -> None:
        record = ExperimentRecord(
            StrategyDefinition("leak_probe", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}),
            {"dataset_id": "fixture", "start_date": "2020-01-01", "end_date": "2021-12-31"}, {},
        )
        calls = []
        def spy(strategy, snapshot, execution):
            calls.append({"fast": strategy["rules"]["fast"], "start": snapshot["start_date"], "end": snapshot["end_date"], "head": snapshot.get("exclude_head_rows", 0), "tail": snapshot.get("exclude_tail_rows", 0)})
            return {"total_return_pct": strategy["rules"]["fast"], "max_drawdown_pct": -1, "trade_count": 5}
        result = optimized_walk_forward(record, spy, folds=2, fast_values=[3, 4], slow_values=[8])
        self.assertFalse(result["summary"]["temporal_leakage_detected"])
        self.assertEqual(len(calls), 6)
        for fold, window in enumerate(result["windows"]):
            selection_calls = calls[fold * 3:fold * 3 + 2]
            test_call = calls[fold * 3 + 2]
            self.assertTrue(all(call["end"] == window["train_end"] for call in selection_calls))
            self.assertEqual(test_call["start"], window["test_start"])
            self.assertEqual(window["selected"]["fast"], 4)

        calls.clear()
        purged = optimized_walk_forward(
            record, spy, folds=2, fast_values=[3, 4], slow_values=[8], purge_days=5, embargo_days=7, purge_rows=3, embargo_rows=4,
        )
        self.assertEqual(purged["purge_days"], 5)
        self.assertEqual(purged["embargo_days"], 7)
        self.assertEqual(purged["purge_rows"], 3)
        self.assertEqual(purged["embargo_rows"], 4)
        for fold in range(2):
            self.assertTrue(all(call["tail"] == 3 for call in calls[fold * 3:fold * 3 + 2]))
            self.assertEqual(calls[fold * 3 + 2]["head"], 4)
        for window in purged["windows"]:
            self.assertEqual(window["excluded_boundary_days"], 12)
            self.assertEqual(window["purge_days"], 5)
            self.assertEqual(window["embargo_days"], 7)
            self.assertLess(window["train_end"], window["test_start"])
        with self.assertRaisesRegex(ValueError, "between 0 and 365"):
            optimized_walk_forward(record, spy, folds=2, fast_values=[3], slow_values=[8], purge_days=366)
        with self.assertRaisesRegex(ValueError, "between 0 and 10000"):
            optimized_walk_forward(record, spy, folds=2, fast_values=[3], slow_values=[8], purge_rows=-1)
        labelled = ExperimentRecord(
            StrategyDefinition("labelled", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8, "label_horizon_rows": 6}),
            record.data_snapshot, {},
        )
        calls.clear(); horizon = optimized_walk_forward(labelled, spy, folds=2, fast_values=[3], slow_values=[8], purge_rows=2)
        self.assertEqual(horizon["requested_purge_rows"], 2)
        self.assertEqual(horizon["declared_label_horizon_rows"], 6)
        self.assertEqual(horizon["purge_rows"], 6)
        self.assertTrue(all(call["tail"] == 6 for call in calls[::2]))

    def test_gateway_create_list_and_run(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            created = call_research_tool(
                "research_experiment_create",
                {
                    "adapter": "mock",
                    "strategy": {"name": "safe", "version": "1", "rules": {"entry": "SMA(5)>SMA(20)"}},
                    "data_snapshot": {"dataset_id": "fixture-v2", "data_mode": "mock"},
                    "execution_model": {"fill_model": "next_bar_open"},
                },
                runtime_root=root,
                repo_root=root,
            )
            experiment_id = created["experiment"]["id"]
            result = call_research_tool(
                "research_backtest_run",
                {"adapter": "mock", "experiment_id": experiment_id},
                runtime_root=root,
                repo_root=root,
            )
            self.assertTrue(result["ok"])
            self.assertEqual(result["experiment"]["status"], "VALIDATION")
            listed = call_research_tool(
                "research_experiment_list", {"adapter": "mock"}, runtime_root=root, repo_root=root
            )
            self.assertEqual(listed["count"], 1)
            second = call_research_tool(
                "research_experiment_create",
                {
                    "adapter": "mock",
                    "strategy": {"name": "safe_two", "version": "1", "rules": {"entry": "SMA(8)>SMA(32)"}},
                    "data_snapshot": {"dataset_id": "fixture-v2", "data_mode": "mock"},
                    "execution_model": {"fill_model": "next_bar_open"},
                },
                runtime_root=root,
                repo_root=root,
            )
            comparison = call_research_tool(
                "research_experiment_compare",
                {"adapter": "mock", "experiment_ids": [experiment_id, second["experiment"]["id"]]},
                runtime_root=root,
                repo_root=root,
            )
            self.assertTrue(comparison["comparison"]["comparable"])

    def test_async_submit_rejects_invalid_collection_before_persisting_job(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            invalid_payloads = (
                ({"symbols": ["005930"]}, "requires start and end"),
                (
                    {"symbols": ["005930"], "start": "2024/01/01", "end": "2024-01-10"},
                    "must be ISO dates",
                ),
                (
                    {"symbols": ["005930"], "start": "2024-02-01", "end": "2024-01-10"},
                    "precedes start date",
                ),
            )
            for payload, message in invalid_payloads:
                with self.subTest(message=message), self.assertRaisesRegex(ValueError, message):
                    call_research_tool(
                        "research_async_submit",
                        {"adapter": "mock", "job_type": "collection", "payload": payload},
                        runtime_root=root,
                        repo_root=root,
                    )

            status = call_research_tool(
                "research_async_status",
                {"adapter": "mock"},
                runtime_root=root,
                repo_root=root,
            )
            self.assertEqual(0, status["job_count"])

    def test_async_submit_accepts_valid_collection_after_preflight(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            submitted = call_research_tool(
                "research_async_submit",
                {
                    "adapter": "mock",
                    "job_type": "collection",
                    "payload": {
                        "provider": "mock",
                        "symbols": ["005930", "005930"],
                        "start": "2024-01-01",
                        "end": "2024-01-10",
                        "timeframe": "1d",
                    },
                },
                runtime_root=root,
                repo_root=root,
            )
            job_id = submitted["job"]["job_id"]
            final = None
            for _ in range(300):
                final = call_research_tool(
                    "research_async_status",
                    {"adapter": "mock", "job_id": job_id},
                    runtime_root=root,
                    repo_root=root,
                )["job"]
                if final["status"] in {"SUCCEEDED", "FAILED", "CANCELLED", "INTERRUPTED"}:
                    break
                time.sleep(0.01)

            self.assertIsNotNone(final)
            self.assertEqual("SUCCEEDED", final["status"])
            self.assertEqual(["005930"], final["payload"]["symbols"])
            self.assertEqual("2024-01-01", final["payload"]["start"])

    def test_async_submit_rejects_missing_experiment_before_persisting_job(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            with self.assertRaisesRegex(ValueError, "requires a non-empty experiment_id"):
                call_research_tool(
                    "research_async_submit",
                    {"adapter": "mock", "job_type": "backtest", "payload": {}},
                    runtime_root=root,
                    repo_root=root,
                )
            status = call_research_tool(
                "research_async_status",
                {"adapter": "mock"},
                runtime_root=root,
                repo_root=root,
            )
            self.assertEqual(0, status["job_count"])

    def test_mcp_server_imports_and_declares_every_research_gateway_tool(self) -> None:
        from codexstock_mcp_server import TOOLS
        from codexstock_research_forge.gateway import RESEARCH_TOOL_NAMES

        declared = {row["name"] for row in TOOLS}
        self.assertTrue(set(RESEARCH_TOOL_NAMES).issubset(declared))

    def test_local_ohlcv_uses_prior_close_signal_and_next_open_fill(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            rows = []
            for index in range(400):
                price = 100.0 + index
                day = date(2023, 1, 1) + timedelta(days=index)
                rows.append(
                    {
                        "date": day.isoformat(),
                        "open": price,
                        "high": price + 1,
                        "low": price - 1,
                        "close": price + 0.5,
                        "volume": 1000 + index,
                    }
                )
            (root / "fixture.json").write_text(
                json.dumps({
                    "005930": {"ok": True, "yahoo": "005930.KS", "rows": rows},
                    "000660": {"ok": True, "yahoo": "000660.KS", "rows": [{**row, "open": row["open"] * 2, "high": row["high"] * 2, "low": row["low"] * 2, "close": row["close"] * 2} for row in rows]},
                }),
                encoding="utf-8",
            )
            forge = ResearchForge.local(root / "runtime", LocalOhlcvBacktestAdapter(root))
            forge.universe().register(
                "fixture-universe-v1",
                [
                    {"symbol": "005930", "name": "Samsung", "market": "KOSPI", "security_type": "COMMON", "listing_date": "1975-01-01"},
                    {"symbol": "000660", "name": "SK Hynix", "market": "KOSPI", "security_type": "COMMON", "listing_date": "1996-01-01"},
                ],
                "test-fixture",
            )
            record = forge.create_experiment(
                StrategyDefinition("real_rows", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}),
                {
                    "dataset_id": "fixture-ohlcv-v1",
                    "data_mode": "historical_provider",
                    "cache_file": "fixture.json",
                    "start_date": "2023-01-01",
                    "end_date": "2024-02-04",
                    "point_in_time_universe": True,
                    "universe_dataset_id": "fixture-universe-v1",
                    "adjusted_prices": True,
                },
                {"initial_cash": 1_000_000, "commission_bps": 1.5, "slippage_bps": 2, "sell_tax_bps": 18, "max_volume_participation": 0.1, "market_impact_bps_at_full_participation": 10, "order_delay_bars": 1},
            )
            completed = forge.run_backtest(record.id)
            self.assertEqual(completed.status, ExperimentStatus.VALIDATION)
            self.assertTrue(completed.validation["passed"])
            self.assertEqual(completed.result["fill_timing"], "open_t_plus_1")
            self.assertEqual(completed.result["data_mode"], "historical_provider")
            walked = forge.run_walk_forward(record.id, folds=4)
            self.assertEqual(walked.validation["walk_forward"]["summary"]["fold_count"], 4)
            strict = forge.run_strict_walk_forward(record.id, folds=2, fast_values=[3, 4], slow_values=[8, 10], purge_rows=2, embargo_rows=1)
            strict_result = strict.validation["strict_walk_forward"]
            self.assertFalse(strict_result["parameter_selection_uses_oos"])
            self.assertFalse(strict_result["summary"]["temporal_leakage_detected"])
            self.assertEqual(strict_result["summary"]["fold_count"], 2)
            for window in strict_result["windows"]:
                self.assertLess(window["train_end"], window["test_start"])
            robust = forge.run_robustness(record.id, radius=1)
            self.assertGreaterEqual(robust.validation["parameter_robustness"]["summary"]["candidate_count"], 6)
            replay = forge.replay(record.id, limit=10)
            self.assertTrue(replay["read_only"])
            report = forge.export(record.id)
            self.assertTrue(Path(report["json_path"]).is_file())
            self.assertTrue(Path(report["markdown_path"]).is_file())
            self.assertTrue(Path(report["excel_path"]).is_file())
            self.assertTrue(Path(report["csv_path"]).is_file())
            self.assertTrue(Path(report["manifest_path"]).is_file())
            self.assertEqual(report["formats"], ["json", "markdown", "xlsx", "csv"])
            self.assertIn("purge_rows=2", Path(report["markdown_path"]).read_text(encoding="utf-8"))
            self.assertIn("strict_walk_forward,purge_rows,2", Path(report["csv_path"]).read_text(encoding="utf-8"))
            from openpyxl import load_workbook
            workbook = load_workbook(report["excel_path"], read_only=True)
            self.assertIn("Strict Walk Forward", workbook.sheetnames)
            self.assertEqual(workbook["Strict Walk Forward"]["H2"].value, 2)
            workbook.close()
            for key in ("json_path", "markdown_path", "excel_path", "csv_path"):
                path = Path(report[key])
                self.assertEqual(hashlib.sha256(path.read_bytes()).hexdigest(), report["artifact_hashes"][path.name]["sha256"])
            self.assertTrue(forge.verify_export(record.id)["ok"])
            Path(report["markdown_path"]).write_text("tampered", encoding="utf-8")
            tampered = forge.verify_export(record.id)
            self.assertFalse(tampered["ok"])
            self.assertTrue(any(value.startswith("artifact_hash_mismatch:") for value in tampered["errors"]))

            portfolio = forge.create_experiment(
                StrategyDefinition(
                    "portfolio",
                    "1",
                    {
                        "type": "portfolio_ma_cross",
                        "symbols": ["005930", "000660"],
                        "fast": 3,
                        "slow": 8,
                        "max_positions": 1,
                        "position_size": 300_000,
                    },
                ),
                {
                    "dataset_id": "fixture-ohlcv-v1",
                    "data_mode": "historical_provider",
                    "cache_file": "fixture.json",
                    "start_date": "2023-01-01",
                    "end_date": "2024-02-04",
                    "point_in_time_universe": True,
                    "universe_dataset_id": "fixture-universe-v1",
                    "adjusted_prices": True,
                },
                {"initial_cash": 1_000_000, "commission_bps": 1.5, "slippage_bps": 2, "sell_tax_bps": 18, "max_volume_participation": 0.1, "market_impact_bps_at_full_participation": 10, "order_delay_bars": 1},
            )
            portfolio = forge.run_backtest(portfolio.id)
            self.assertLessEqual(max(row["positions"] for row in portfolio.result["equity_curve"]), 1)
            self.assertEqual(portfolio.result["entry_priority"], "symbol_ascending_deterministic")
            self.assertGreater(portfolio.result["partial_fill_count"], 0)
            self.assertTrue(any(row["reason"] == "MAX_POSITIONS" for row in portfolio.result["rejected_orders"]))

            indicator_strategy = forge.create_experiment(
                StrategyDefinition(
                    "indicator_combo", "1",
                    {
                        "type": "indicator_rules", "symbol": "005930", "profile": "STANDARD",
                        "entry": {"all": [{"left": {"field": "close"}, "operator": ">", "right": {"indicator": "SMA", "parameters": {"period": 5}, "output": "value"}}]},
                        "exit": {"all": [{"left": {"field": "close"}, "operator": "<", "right": {"indicator": "SMA", "parameters": {"period": 5}, "output": "value"}}]},
                    },
                ),
                {
                    "dataset_id": "fixture-ohlcv-v1", "data_mode": "historical_provider", "cache_file": "fixture.json",
                    "start_date": "2023-01-01", "end_date": "2024-02-04", "universe_dataset_id": "fixture-universe-v1", "adjusted_prices": True,
                },
                {"initial_cash": 1_000_000, "commission_bps": 1.5, "slippage_bps": 2, "sell_tax_bps": 18, "max_volume_participation": 0.1, "market_impact_bps_at_full_participation": 10, "order_delay_bars": 1},
            )
            indicator_strategy = forge.run_backtest(indicator_strategy.id)
            self.assertGreater(indicator_strategy.result["entry_signal_count"], 0)
            self.assertEqual(indicator_strategy.result["trades"][0]["side"], "BUY")
            self.assertLess(indicator_strategy.result["trades"][0]["signal_date"], indicator_strategy.result["trades"][0]["date"])

    def test_universe_requires_registered_interval_not_boolean_claim(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = UniverseStore(root / "universes")
            meta = store.register(
                "krx-history-v1",
                [{"symbol": "005930", "market": "KOSPI", "security_type": "COMMON", "listing_date": "1975-01-01", "delisting_date": None}],
                "KRX fixture",
            )
            self.assertTrue(meta["content_hash"].startswith("sha256:"))
            self.assertEqual(store.query("krx-history-v1", "2020-01-01")["count"], 1)
            self.assertTrue(store.validate_period("krx-history-v1", ["005930"], "2019-01-01", "2023-12-31")["passed"])
            self.assertFalse(store.validate_period("krx-history-v1", ["999999"], "2019-01-01", "2023-12-31")["passed"])
            forged = validate_experiment_evidence(
                {"point_in_time_universe": True, "adjusted_prices": True},
                {"commission_bps": 1, "slippage_bps": 1, "sell_tax_bps": 1},
                {"data_mode": "historical_provider", "dataset_hash": "sha256:x", "data_quality": {"strict_temporal_order": True, "invalid_ohlc_rows": 0}, "signal_timing": "close_t", "fill_timing": "open_t_plus_1"},
                None,
            )
            self.assertIn("point_in_time_universe", forged["blockers"])

    def test_krx_snapshot_is_official_but_not_multi_day_history(self) -> None:
        class Response:
            def __enter__(self): return self
            def __exit__(self, *args): return False
            def read(self):
                return json.dumps({"OutBlock_1": [{"ISU_SRT_CD": "005930", "ISU_ABBRV": "Samsung", "LIST_DD": "19750611", "SECUGRP_NM": "주권"}]}).encode()

        provider = KrxOpenApiUniverseProvider("secret-key")
        with patch("urllib.request.urlopen", return_value=Response()):
            snapshot = provider.fetch_snapshot("2024-01-02", ["KOSPI"])
        self.assertEqual(snapshot["records"][0]["symbol"], "005930")
        self.assertEqual(snapshot["evidence"]["grade"], "official_snapshot")
        with tempfile.TemporaryDirectory() as directory:
            store = UniverseStore(Path(directory))
            store.register("krx-20240102", snapshot["records"], snapshot["source"], snapshot["evidence"])
            self.assertTrue(store.validate_period("krx-20240102", ["005930"], "2024-01-02", "2024-01-02")["passed"])
            evidence = store.validate_period("krx-20240102", ["005930"], "2024-01-01", "2024-01-03")
            self.assertFalse(evidence["passed"])
            self.assertIn("dataset_coverage", evidence["uncovered_symbols"])

    def test_krx_global_history_restores_delisted_members_without_survivorship_bias(self) -> None:
        snapshot = {"as_of": "2024-12-31", "records": [{"symbol": "005930", "name": "Samsung", "market": "KOSPI", "security_type": "COMMON", "listing_date": "1975-06-11", "delisting_date": None}]}
        listings = [{"isu_cd": "123456", "co_nm": "NewCo", "lst_dt": "2024/03/02"}]
        delistings = [
            {"isu_cd": "123456", "kor_cor_nm": "NewCo", "chg_dt": "2024/09/10", "tr_stp_rsn": "DELISTING (KOSDAQ)"},
            {"isu_cd": "654321", "kor_cor_nm": "OldCo", "chg_dt": "2024/07/01", "tr_stp_rsn": "DELISTING (Stock)"},
        ]
        provider = KrxGlobalListingHistoryProvider()
        with patch.object(provider, "_query_retry", side_effect=[(listings, b"listing"), ([], b"relisting"), (delistings, b"delisting")]):
            history = provider.fetch_history("2024-01-01", "2024-12-31", snapshot)
        self.assertEqual(history["delisted_record_count"], 2)
        self.assertEqual(history["clipped_precoverage_listing_count"], 1)
        self.assertTrue(history["evidence"]["includes_delisted"])
        self.assertFalse(history["evidence"]["complete_daily_history"])
        with tempfile.TemporaryDirectory() as directory:
            store = UniverseStore(Path(directory)); store.register("official-history", history["records"], history["source"], history["evidence"])
            april = {row["symbol"] for row in store.query("official-history", "2024-04-01")["symbols"]}
            october = {row["symbol"] for row in store.query("official-history", "2024-10-01")["symbols"]}
            self.assertEqual(april, {"005930", "123456", "654321"})
            self.assertEqual(october, {"005930"})

    def test_kind_snapshot_parser_and_history_survivorship_events(self) -> None:
        today = date.today().isoformat()
        rows = []
        for index in range(1001):
            rows.append(f"<tr><td>Company {index}</td><td>유가</td><td>{index:06d}</td><td>Industry</td><td>Product</td><td>2020-01-02</td><td>12월</td><td>CEO</td><td></td><td>서울</td></tr>")
        html = ("<table><tr><th>회사명</th><th>시장구분</th><th>종목코드</th><th>업종</th><th>주요제품</th><th>상장일</th><th>결산월</th><th>대표자명</th><th>홈페이지</th><th>지역</th></tr>" + "".join(rows) + "</table>").encode("euc-kr")
        class Response:
            headers = {"Content-Type": "application/vnd.ms-excel; charset=EUC-KR"}
            def __enter__(self): return self
            def __exit__(self, *args): return False
            def read(self): return html
        with patch("codexstock_research_forge.universe.urllib.request.urlopen", return_value=Response()):
            snapshot = KindOfficialUniverseProvider().fetch_snapshot(today)
        self.assertEqual(len(snapshot["records"]), 1001)
        self.assertTrue(snapshot["evidence"]["official"])
        self.assertFalse(snapshot["evidence"]["includes_delisted"])

        with tempfile.TemporaryDirectory() as directory:
            history = PointInTimeUniverseHistory(Path(directory))
            baseline = {**snapshot, "records": snapshot["records"][:2]}
            created = history.create_baseline("official-chain", baseline)
            self.assertEqual(created["snapshot_count"], 1)
            next_snapshot = {**baseline, "as_of": (date.today() + timedelta(days=1)).isoformat(), "records": [snapshot["records"][1], snapshot["records"][2]], "source_hash": "sha256:next"}
            appended = history.append_snapshot("official-chain", next_snapshot)
            self.assertEqual(appended["new_event_count"], 2)
            before = history.query("official-chain", today)
            after = history.query("official-chain", next_snapshot["as_of"])
            self.assertEqual({row["symbol"] for row in before["symbols"]}, {"000000", "000001"})
            self.assertEqual({row["symbol"] for row in after["symbols"]}, {"000001", "000002"})
            self.assertNotEqual(before["result_hash"], after["result_hash"])
            changed = history.record_official_code_change(
                "official-chain", next_snapshot["as_of"], "000001",
                {**snapshot["records"][3], "listing_date": next_snapshot["as_of"]},
                "https://kind.krx.co.kr/corporate-action", "sha256:" + "0" * 64,
            )
            self.assertFalse(changed["idempotent"])
            changed_state = history.query("official-chain", next_snapshot["as_of"])
            self.assertEqual({row["symbol"] for row in changed_state["symbols"]}, {"000002", "000003"})
            with self.assertRaisesRegex(ValueError, "outside"):
                history.query("official-chain", "2019-01-01")

    def test_collection_checkpoint_partial_failure_and_resume(self) -> None:
        class FlakyProvider(MockMarketDataProvider):
            def fetch_daily_bars(self, symbol: str, start: str, end: str):
                if symbol == "000660":
                    raise RuntimeError("temporary provider failure")
                return super().fetch_daily_bars(symbol, start, end)

        with tempfile.TemporaryDirectory() as directory:
            manager = CollectionManager(Path(directory))
            partial = manager.start(FlakyProvider(), ["005930", "000660"], "1d", "2024-01-01", "2024-01-10")
            self.assertFalse(partial["ok"])
            self.assertEqual(partial["job"]["status"], "PARTIAL")
            self.assertEqual(partial["job"]["completed_symbols"], ["005930"])
            resumed = manager.resume(partial["job"]["job_id"], MockMarketDataProvider())
            self.assertTrue(resumed["ok"])
            self.assertEqual(resumed["job"]["retry_count"], 1)
            self.assertEqual(resumed["job"]["completed_symbols"], ["000660", "005930"])
            self.assertEqual(manager.storage_summary()["bar_files"], 2)

    def test_universe_integrity_rejects_overlaps_and_requires_lineage_for_complete_history(self) -> None:
        old = {"symbol": "000001", "name": "Old Issuer", "market": "KOSPI", "security_type": "COMMON", "listing_date": "2010-01-01", "delisting_date": "2015-12-31"}
        new = {"symbol": "000001", "name": "New Issuer", "market": "KOSPI", "security_type": "COMMON", "listing_date": "2016-01-04"}
        with tempfile.TemporaryDirectory() as directory:
            store = UniverseStore(Path(directory))
            store.register("reuse-incomplete", [old, new], "fixture", {"grade": "declared_intervals", "complete_daily_history": False})
            audit = store.integrity("reuse-incomplete")
            self.assertEqual(audit["code_reuse_transition_count"], 1)
            self.assertEqual(audit["missing_lineage_evidence_count"], 1)
            self.assertFalse(audit["lineage_evidence_complete"])
            with self.assertRaisesRegex(ValueError, "lineage evidence"):
                store.register("reuse-claimed-complete", [old, new], "fixture", {"grade": "complete_listing_history", "complete_daily_history": True})
            source = {"source_url": "https://kind.krx.co.kr/lineage", "source_hash": "sha256:" + "f" * 64}
            verified = [{**old, **source, "lineage_id": "issuer-old"}, {**new, **source, "lineage_id": "issuer-new"}]
            store.register("reuse-verified", verified, "official", {"grade": "complete_listing_history", "complete_daily_history": True})
            self.assertTrue(store.integrity("reuse-verified")["lineage_evidence_complete"])
            overlap = {**new, "listing_date": "2015-12-31"}
            with self.assertRaisesRegex(ValueError, "overlapping"):
                store.register("overlap", [old, overlap], "fixture")

    def test_universe_store_detects_persisted_content_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            store = UniverseStore(root)
            store.register("tamper-check", [{
                "symbol": "005930", "market": "KOSPI", "security_type": "STOCK",
                "listing_date": "1975-06-11", "delisting_date": None,
            }], "fixture")
            path = root / "tamper-check.json"
            payload = json.loads(path.read_text(encoding="utf-8"))
            payload["records"][0]["listing_date"] = "1975-06-12"
            path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "integrity"):
                store.get("tamper-check")
            status = store.status()
            self.assertFalse(status["ok"])
            self.assertEqual(status["dataset_count"], 0)
            self.assertEqual(len(status["invalid"]), 1)

    def test_collection_retries_rate_limit_and_network_faults_but_not_permanent_errors(self) -> None:
        class FaultProvider(MockMarketDataProvider):
            def __init__(self): self.calls = {}
            def fetch_daily_bars(self, symbol: str, start: str, end: str):
                count = self.calls.get(symbol, 0) + 1; self.calls[symbol] = count
                if symbol == "005930" and count == 1: raise RuntimeError("HTTP 429 rate limit")
                if symbol == "000660" and count == 1: raise TimeoutError("network timed out")
                if symbol == "035420": raise ValueError("invalid permanent request")
                return super().fetch_daily_bars(symbol, start, end)

        with tempfile.TemporaryDirectory() as directory:
            provider = FaultProvider(); manager = CollectionManager(Path(directory))
            result = manager.start(provider, ["005930", "000660", "035420"], "1d", "2024-01-01", "2024-01-10", retry_max_attempts=3, retry_base_seconds=0)
            job = result["job"]
            self.assertEqual(job["completed_symbols"], ["000660", "005930"])
            self.assertEqual(job["failed_symbols"]["035420"]["classification"], "PERMANENT_PROVIDER_ERROR")
            self.assertEqual(provider.calls, {"005930": 2, "000660": 2, "035420": 1})
            self.assertEqual(job["transient_retry_count"], 2)
            self.assertEqual([row["category"] for row in job["retry_events"]], ["RATE_LIMIT", "TRANSIENT_NETWORK", "PERMANENT_PROVIDER_ERROR"])
            self.assertTrue(all("at" in row for row in job["retry_events"]))

    def test_duckdb_parquet_storage_deduplicates_and_queries(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manager = CollectionManager(root / "collection")
            manager.start(MockMarketDataProvider(), ["005930", "000660"], "1d", "2024-01-01", "2024-01-10")
            storage = AnalyticalStorage(root / "analytical")
            first = storage.import_collection(manager.root)
            second = storage.import_collection(manager.root)
            self.assertGreater(first["inserted_net"], 0)
            self.assertEqual(second["inserted_net"], 0)
            result = storage.query(["005930"], "2024-01-01T00:00:00+00:00", "2024-01-31T23:59:59+00:00", "1d")
            self.assertGreater(result["count"], 0)
            self.assertTrue(result["result_hash"].startswith("sha256:"))
            exported = storage.export_parquet()
            self.assertGreater(exported["parquet_files"], 0)
            self.assertEqual(exported["partitioning"], ["timeframe", "year", "month", "symbol_bucket"])

    def test_storage_streams_legacy_top_level_symbol_cache(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            legacy = root / "legacy.json"
            legacy.write_text(json.dumps({
                "005930": {"version": "adj_v2", "rows": [{"date": "2024-01-02", "open": 10, "high": 12, "low": 9, "close": 11, "volume": 100}]},
                "000660": {"version": "adj_v2", "rows": [{"date": "2024-01-02", "open": 20, "high": 22, "low": 19, "close": 21, "volume": 200}]},
            }), encoding="utf-8")
            storage = AnalyticalStorage(root / "analytical")
            result = storage.import_legacy_ohlcv(legacy, ["005930"])
            self.assertTrue(result["streaming"])
            self.assertEqual(result["imported_symbols"], 1)
            self.assertEqual(result["row_count"], 1)
            resumed = storage.import_legacy_ohlcv(legacy, ["005930"])
            self.assertEqual(resumed["imported_symbols"], 0)
            self.assertEqual(resumed["received"], 0)
            self.assertEqual(resumed["resumed_completed_symbols"], 1)

    def test_storage_transaction_rolls_back_malformed_batch(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            storage = AnalyticalStorage(Path(directory))
            valid = {"symbol": "005930", "timestamp": "2024-01-02T00:00:00+09:00", "open": 10, "high": 11, "low": 9, "close": 10, "volume": 100, "source": "fixture"}
            storage.ingest([valid], "1d")
            malformed = {**valid, "symbol": "000660", "timestamp": "not-a-timestamp"}
            with self.assertRaises(Exception):
                storage.ingest([{**valid, "symbol": "035420", "timestamp": "2024-01-03T00:00:00+09:00"}, malformed], "1d")
            self.assertEqual(storage.summary()["row_count"], 1)

    def test_storage_bulk_path_preserves_explicit_timezone(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            storage = AnalyticalStorage(Path(directory))
            rows = [{"symbol": f"{index:06d}", "timestamp": "2024-01-02T00:00:00+09:00", "open": 10, "high": 11, "low": 9, "close": 10, "volume": 100, "source": "bulk-fixture"} for index in range(500)]
            storage.ingest(rows, "1d")
            result = storage.query(["000000"], "2024-01-01T00:00:00+09:00", "2024-01-03T00:00:00+09:00", "1d")
            self.assertEqual(result["rows"][0]["timestamp"], "2024-01-02T00:00:00+09:00")

    def test_performance_gate_writes_machine_readable_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            storage = AnalyticalStorage(root / "storage")
            storage.ingest([{"symbol": "005930", "timestamp": "2024-01-02T00:00:00+09:00", "open": 10, "high": 11, "low": 9, "close": 10, "volume": 100, "source": "fixture"}], "1d")
            micro = MicrostructureStore(root / "microstructure")
            micro.ingest([{"event_type": "tick", "symbol": "005930", "timestamp": f"2024-01-02T09:00:{index:02d}+09:00", "source": "fixture", "payload": {"price": 100 + index, "quantity": 1}} for index in range(20)])
            archive = MicrostructureArchive(root / "microstructure"); archive.export_incremental()
            result = run_performance_benchmark(storage, root / "benchmarks", 1000, archive)
            self.assertTrue(result["passed"])
            self.assertEqual({row["name"] for row in result["measurements"]}, {"duckdb_bounded_query", "all_indicators", "microstructure_archive_concurrent_reads"})
            archive_measurement = next(row for row in result["measurements"] if row["name"] == "microstructure_archive_concurrent_reads")
            self.assertEqual(archive_measurement["unique_count_pairs"], [[20, 20]])
            self.assertEqual(len(archive_measurement["unique_result_hashes"]), 1)
            self.assertTrue(result["evidence_hash"].startswith("sha256:"))
            self.assertTrue(Path(result["path"]).is_file())
            evidence_path = Path(result["path"]); tampered = json.loads(evidence_path.read_text(encoding="utf-8")); tampered["generated_at"] = "tampered"
            evidence_path.write_text(json.dumps(tampered), encoding="utf-8")
            self.assertFalse(_verified_performance(evidence_path)["passed"])
            minimal = {"schema_version": 1, "passed": True, "measurements": []}; canonical = json.dumps(minimal, ensure_ascii=False, sort_keys=True, separators=(",", ":")); minimal["evidence_hash"] = "sha256:" + hashlib.sha256(canonical.encode()).hexdigest()
            evidence_path.write_text(json.dumps(minimal), encoding="utf-8")
            self.assertFalse(_verified_performance(evidence_path)["passed"])

    def test_concurrent_storage_collection_and_indicator_soak_cleans_up(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); storage = AnalyticalStorage(root / "storage")
            result = run_concurrency_soak(storage, CollectionManager(root / "collection"), root / "benchmarks", iterations=2)
            self.assertTrue(result["passed"], result["errors"])
            self.assertTrue(result["evidence_hash"].startswith("sha256:"))
            self.assertEqual(result["measurements"]["storage_write"]["rows"], 200)
            self.assertEqual(storage.summary()["row_count"], 0)
            evidence_path = Path(result["path"]); altered = json.loads(evidence_path.read_text(encoding="utf-8")); altered["elapsed_seconds"] = 0
            evidence_path.write_text(json.dumps(altered), encoding="utf-8")
            self.assertFalse(_verified_performance(evidence_path)["passed"])

    def test_async_jobs_persist_progress_cancel_failure_and_retry(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            manager = AsyncJobManager(Path(directory), max_workers=1)

            def successful(payload, progress, cancelled):
                progress(40, "working")
                return {"value": payload["value"]}

            submitted = manager.submit("backtest", {"value": 7}, successful)
            completed = self._wait_job(manager, submitted["job_id"])
            self.assertEqual(completed["status"], "SUCCEEDED")
            self.assertEqual(completed["progress_percent"], 100.0)
            self.assertEqual(completed["result"]["value"], 7)

            def cooperative(payload, progress, cancelled):
                for index in range(100):
                    if cancelled():
                        return {"stopped_at": index}
                    progress(index, "loop")
                    time.sleep(0.005)
                return {}

            running = manager.submit("collection", {}, cooperative)
            time.sleep(0.03)
            manager.cancel(running["job_id"])
            cancelled = self._wait_job(manager, running["job_id"])
            self.assertEqual(cancelled["status"], "CANCELLED")

            attempts = {"count": 0}
            def flaky(payload, progress, cancelled):
                attempts["count"] += 1
                if attempts["count"] == 1:
                    raise RuntimeError("injected failure")
                return {"recovered": True}

            failed_job = manager.submit("backtest", {}, flaky)
            failed = self._wait_job(manager, failed_job["job_id"])
            self.assertEqual(failed["status"], "FAILED")
            retried = manager.retry(failed["job_id"], flaky)
            recovered = self._wait_job(manager, retried["job_id"])
            self.assertEqual(recovered["status"], "SUCCEEDED")
            self.assertEqual(recovered["attempt"], 2)

    def test_async_cancellation_wins_over_handler_exception(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            manager = AsyncJobManager(Path(directory), max_workers=1)
            entered = threading.Event()

            def aborting(payload, progress, cancelled):
                entered.set()
                while not cancelled():
                    time.sleep(0.001)
                raise RuntimeError("handler aborted because cancellation was requested")

            submitted = manager.submit("collection", {}, aborting)
            self.assertTrue(entered.wait(1.0))
            manager.cancel(submitted["job_id"])
            cancelled = self._wait_job(manager, submitted["job_id"])
            self.assertEqual(cancelled["status"], "CANCELLED")
            self.assertIsNone(cancelled["error"])

    def test_async_status_polling_is_safe_during_atomic_progress_writes(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            manager = AsyncJobManager(Path(directory), max_workers=1)

            def busy(payload, progress, cancelled):
                for index in range(300):
                    progress(index / 3.1, "rapid-progress")
                return {"writes": 300}

            submitted = manager.submit("collection", {}, busy)
            polls = 0
            while True:
                snapshot = manager.status()
                polls += 1
                matching = [row for row in snapshot["jobs"] if row["job_id"] == submitted["job_id"]]
                self.assertEqual(len(matching), 1)
                if matching[0]["status"] in {"SUCCEEDED", "FAILED", "CANCELLED"}:
                    final = matching[0]
                    break
            self.assertEqual(final["status"], "SUCCEEDED")
            self.assertEqual(final["result"]["writes"], 300)
            self.assertGreater(polls, 0)

    def test_async_manager_marks_orphaned_running_job_interrupted(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            job_id = "job_" + "a" * 32
            (root / f"{job_id}.json").write_text(json.dumps({
                "schema_version": 1, "job_id": job_id, "type": "backtest",
                "status": "RUNNING", "progress_percent": 55, "phase": "backtest_running",
                "payload": {}, "result": None, "error": None, "cancel_requested": False,
                "attempt": 1, "created_at": "2024-01-01T00:00:00+00:00",
                "updated_at": "2024-01-01T00:00:00+00:00", "started_at": "2024-01-01T00:00:00+00:00", "finished_at": None,
            }), encoding="utf-8")
            recovered = AsyncJobManager(root).get(job_id)
            self.assertEqual(recovered["status"], "INTERRUPTED")
            self.assertEqual(recovered["error"]["type"], "WorkerOwnerLost")

    def test_async_worker_atomically_resumes_interrupted_jobs_once(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            manager = AsyncJobManager(root, max_workers=1)

            def interrupted(job_id: str) -> None:
                (root / f"{job_id}.json").write_text(json.dumps({
                    "schema_version": 1, "job_id": job_id, "type": "backtest",
                    "status": "INTERRUPTED", "progress_percent": 55, "phase": "process_restarted",
                    "payload": {"value": 11}, "result": None,
                    "error": {"type": "ProcessRestart", "message": "fixture"}, "cancel_requested": False,
                    "attempt": 1, "created_at": "2024-01-01T00:00:00+00:00",
                    "updated_at": "2024-01-01T00:00:00+00:00", "started_at": None, "finished_at": None,
                }), encoding="utf-8")

            first_id = "job_" + "b" * 32
            interrupted(first_id)
            calls = {"count": 0}
            def resolver(job_type):
                self.assertEqual(job_type, "backtest")
                def handler(payload, progress, cancelled):
                    calls["count"] += 1
                    progress(80, "resumed")
                    return {"value": payload["value"]}
                return handler

            recovery = manager.resume_interrupted(resolver)
            self.assertEqual(recovery["resumed_job_ids"], [first_id])
            completed = self._wait_job(manager, first_id)
            self.assertEqual(completed["status"], "SUCCEEDED")
            self.assertEqual(completed["attempt"], 2)
            self.assertEqual(completed["resume_count"], 1)
            self.assertEqual(calls["count"], 1)
            self.assertEqual(manager.resume_interrupted(resolver)["resumed_count"], 0)

            claimed_id = "job_" + "c" * 32
            interrupted(claimed_id)
            claimed_lease = root / f"{claimed_id}.lease"
            claimed_lease.mkdir()
            (claimed_lease / "owner.json").write_text(json.dumps({"pid": os.getpid()}), encoding="utf-8")
            conflict = manager.resume_interrupted(resolver)
            self.assertEqual(conflict["resumed_count"], 0)
            self.assertEqual(conflict["claim_conflicts"], [claimed_id])

            (claimed_lease / "owner.json").write_text(json.dumps({"pid": 99999999}), encoding="utf-8")
            reclaimed = manager.resume_interrupted(resolver)
            self.assertEqual(reclaimed["resumed_job_ids"], [claimed_id])
            self.assertEqual(self._wait_job(manager, claimed_id)["status"], "SUCCEEDED")

    def test_async_orphan_detector_preserves_live_owner_and_recovers_dead_owner(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); manager = AsyncJobManager(root)
            job_id = "job_" + "d" * 32
            (root / f"{job_id}.json").write_text(json.dumps({
                "schema_version": 1, "job_id": job_id, "type": "backtest", "status": "RUNNING",
                "progress_percent": 25, "phase": "running", "payload": {}, "result": None, "error": None,
                "cancel_requested": False, "attempt": 1, "created_at": "2024-01-01T00:00:00+00:00",
                "updated_at": "2024-01-01T00:00:00+00:00", "started_at": "2024-01-01T00:00:00+00:00", "finished_at": None,
            }), encoding="utf-8")
            lease = root / f"{job_id}.lease"; lease.mkdir()
            owner = lease / "owner.json"; owner.write_text(json.dumps({"pid": os.getpid()}), encoding="utf-8")
            live = manager.recover_orphaned_running(0)
            self.assertEqual(live["active_job_ids"], [job_id])
            self.assertEqual(manager.get(job_id)["status"], "RUNNING")

            owner.write_text(json.dumps({"pid": 99999999}), encoding="utf-8")
            dead = manager.recover_orphaned_running(0)
            self.assertEqual(dead["recovered_job_ids"], [job_id])
            self.assertEqual(manager.get(job_id)["error"]["type"], "WorkerOwnerLost")
            self.assertFalse(lease.exists())

    def test_replay_merges_bars_microstructure_and_trades_with_cursor_state(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            forge = ResearchForge.local(root)
            record = forge.create_experiment(
                StrategyDefinition("replay", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}),
                {"dataset_id": "replay-fixture"}, {"commission_bps": 1, "slippage_bps": 1},
            )
            record.result = {"trades": [{"symbol": "005930", "timestamp": "2024-01-02T09:01:00+09:00", "side": "BUY", "price": 101}]}
            forge.registry.save(record)
            forge.storage().ingest([
                {"symbol": "005930", "timestamp": "2024-01-02T09:00:00+09:00", "open": 100, "high": 102, "low": 99, "close": 101, "volume": 1000, "source": "fixture"},
            ], "1m")
            forge.microstructure().ingest([
                {"event_type": "tick", "symbol": "005930", "timestamp": "2024-01-02T09:00:30+09:00", "source": "fixture", "payload": {"price": 101, "quantity": 3}},
                {"event_type": "orderbook", "symbol": "005930", "timestamp": "2024-01-02T09:00:45+09:00", "source": "fixture", "payload": {"asks": [[102, 10]], "bids": [[100, 12]]}},
            ])
            session = forge.replay_engine().create(record, ["005930"], "2024-01-02T00:00:00+09:00", "2024-01-03T00:00:00+09:00", ["1m"])
            self.assertEqual(session["event_count"], 4)
            self.assertEqual(session["schema_version"], 2)
            self.assertTrue((root / "replays" / session["frames_file"]).is_file())
            self.assertTrue((root / "replays" / session["index_file"]).is_file())
            first = forge.replay_engine().page(session["session_id"], 0, 2)
            second = forge.replay_engine().page(session["session_id"], first["next_cursor"], 2)
            self.assertTrue(first["has_more"])
            self.assertFalse(second["has_more"])
            self.assertIn("005930", second["state"]["last_orderbook"])
            self.assertEqual(first["timeline_hash"], second["timeline_hash"])
            self.assertTrue(second["indexed"])
            self.assertLessEqual(second["frames_scanned"], 1000 + 2)
            verified = forge.replay_engine().verify(session["session_id"])
            self.assertTrue(verified["verified"])
            self.assertEqual(verified["checked_frames"], 4)
            index_path = root / "replays" / session["index_file"]
            original_index = index_path.read_text(encoding="utf-8")
            index_path.write_text(original_index.replace('"cursor": 0', '"cursor": 1'), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "index integrity"):
                forge.replay_engine().page(session["session_id"], 0, 1)
            self.assertIn("index_file_hash_mismatch", forge.replay_engine().verify(session["session_id"])["errors"])
            index_path.write_text(original_index, encoding="utf-8")
            frames_path = root / "replays" / session["frames_file"]
            with frames_path.open("ab") as handle: handle.write(b"{}\n")
            frame_tamper = forge.replay_engine().verify(session["session_id"])
            self.assertFalse(frame_tamper["verified"])
            self.assertIn("frames_file_hash_mismatch", frame_tamper["errors"])

    def test_replay_hybrid_deduplicates_archive_and_live_and_archive_survives_log_removal(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); forge = ResearchForge.local(root)
            record = forge.create_experiment(StrategyDefinition("archive-replay", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}), {"dataset_id": "fixture"}, {})
            events = [
                {"event_type": "tick", "symbol": "005930", "timestamp": "2024-01-02T09:00:00+09:00", "source": "fixture", "payload": {"price": 100, "volume": 1}},
                {"event_type": "orderbook", "symbol": "005930", "timestamp": "2024-01-02T09:00:01+09:00", "source": "fixture", "payload": {"asks": [[101, 2]], "bids": [[99, 3]]}},
            ]
            forge.microstructure().ingest(events); forge.microstructure_archive().export_incremental()
            hybrid = forge.replay_engine().create(record, ["005930"], "2024-01-02T00:00:00+09:00", "2024-01-03T00:00:00+09:00", [], 100, "hybrid")
            self.assertEqual(hybrid["event_count"], 2)
            self.assertEqual(hybrid["microstructure_source_counts"], {"archive": 2, "live": 2})
            self.assertEqual(hybrid["microstructure_unique_count"], 2)
            for path in (root / "microstructure" / "events").rglob("*.jsonl"): path.unlink()
            archived = forge.replay_engine().create(record, ["005930"], "2024-01-02T00:00:00+09:00", "2024-01-03T00:00:00+09:00", [], 100, "archive")
            self.assertEqual(archived["event_count"], 2)
            self.assertEqual(archived["lane_counts"], {"orderbook": 1, "tick": 1})

    @staticmethod
    def _wait_job(manager: AsyncJobManager, job_id: str) -> dict:
        for _ in range(300):
            job = manager.get(job_id)
            if job["status"] in {"SUCCEEDED", "FAILED", "CANCELLED", "INTERRUPTED"}:
                return job
            time.sleep(0.01)
        raise AssertionError("async job did not finish")

    def test_kis_provider_is_read_only_and_normalizes_bridge_rows(self) -> None:
        class FakeBridge:
            def status(self):
                return {"configured": True, "readonly": True, "live_trading": False, "order_allowed": False, "account_masked": "12***34"}

            def daily_chart(self, **kwargs):
                return {"ok": True, "rows": [{"date": "2024-01-02", "open": 100, "high": 110, "low": 90, "close": 105, "volume": 1000}]}

            def minute_chart(self, **kwargs):
                return {"ok": True, "rows": [{"date": "2024-01-02", "datetime": "2024-01-02T09:00:00+09:00", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 10}]}

        provider = KisReadOnlyProvider(FakeBridge(), ["005930"], throttle_seconds=0)
        self.assertFalse(provider.status()["order_allowed"])
        daily = provider.fetch_daily_bars("005930", "2024-01-01", "2024-01-10")
        minute = provider.fetch_minute_bars("005930", 1, "2024-01-02", "2024-01-02")
        self.assertEqual(daily[0]["timestamp"], "2024-01-02T00:00:00+09:00")
        self.assertEqual(minute[0]["timestamp"], "2024-01-02T09:00:00+09:00")
        with self.assertRaisesRegex(ValueError, "raw 1-minute"):
            provider.fetch_minute_bars("005930", 5, "2024-01-02", "2024-01-02")

    def test_execution_partial_fill_limits_suspension_and_delay(self) -> None:
        rows = [
            {"date": "2024-01-01", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 10},
            {"date": "2024-01-02", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 10, "upper_limit_locked": True},
            {"date": "2024-01-03", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 10},
            {"date": "2024-01-04", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 10, "lower_limit_locked": True},
        ]
        result = simulate_long_only(
            rows,
            [True, True, False, False],
            [False, False, True, False],
            {"initial_cash": 100_000, "commission_bps": 0, "slippage_bps": 0, "sell_tax_bps": 0, "max_volume_participation": 0.1, "order_delay_bars": 1},
        )
        self.assertEqual(result["trades"][0]["date"], "2024-01-03")
        self.assertEqual(result["trades"][0]["quantity"], 1)
        self.assertTrue(result["trades"][0]["partial_fill"])
        self.assertEqual(result["rejected_orders"][0]["reason"], "UPPER_LIMIT_BUY_BLOCKED")
        self.assertEqual(result["rejected_orders"][1]["reason"], "LOWER_LIMIT_SELL_BLOCKED")
        self.assertEqual(result["open_position_quantity"], 1)

    def test_execution_rejects_empty_nonfinite_and_invalid_ohlc(self) -> None:
        model = {"commission_bps": 0, "slippage_bps": 0, "sell_tax_bps": 0, "order_delay_bars": 1, "max_volume_participation": 0.1}
        with self.assertRaisesRegex(ValueError, "at least one"):
            simulate_long_only([], [], [], model)
        bad = [{"date": "2024-01-01", "open": float("nan"), "high": 11, "low": 9, "close": 10, "volume": 100}]
        with self.assertRaisesRegex(ValueError, "invalid price"):
            simulate_long_only(bad, [False], [False], model)
        crossed = [{"date": "2024-01-01", "open": 10, "high": 9, "low": 11, "close": 10, "volume": 100}]
        with self.assertRaisesRegex(ValueError, "invalid OHLCV"):
            simulate_long_only(crossed, [False], [False], model)

    def test_execution_presets_queue_position_and_vi_are_deterministic(self) -> None:
        manifest_payload = execution_manifest()
        self.assertEqual(set(manifest_payload["presets"]), {"OPTIMISTIC", "REALISTIC", "CONSERVATIVE"})
        rows = []
        for index in range(8):
            price = 100 + index * 5
            rows.append({"date": f"2024-01-{index + 1:02d}", "open": price, "high": price + 2, "low": price - 2, "close": price + 1, "volume": 1000})
        entry, exit_values = [True] + [False] * 7, [False] * 5 + [True] + [False] * 2
        compared = compare_execution_modes(rows, entry, exit_values, {"initial_cash": 100_000})
        modes = compared["modes"]
        self.assertGreaterEqual(modes["OPTIMISTIC"]["total_return_pct"], modes["REALISTIC"]["total_return_pct"])
        self.assertGreaterEqual(modes["REALISTIC"]["total_return_pct"], modes["CONSERVATIVE"]["total_return_pct"])
        self.assertGreater(modes["OPTIMISTIC"]["filled_quantity"], modes["CONSERVATIVE"]["filled_quantity"])

        queue_rows = [
            {"date": "2024-02-01", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 1000},
            {"date": "2024-02-02", "open": 100, "high": 101, "low": 99, "close": 100, "volume": 1000, "ask_executed_quantity": 50, "ask_queue_ahead": 40},
        ]
        queue_model = {"initial_cash": 100_000, "commission_bps": 0, "slippage_bps": 0, "sell_tax_bps": 0, "max_volume_participation": 1, "market_impact_bps_at_full_participation": 0, "order_delay_bars": 1, "queue_model": "QUEUE_AWARE_IF_AVAILABLE", "queue_ahead_multiplier": 1, "missing_queue_fill_haircut": 1}
        queued = simulate_long_only(queue_rows, [True, False], [False, False], queue_model)
        self.assertEqual(queued["trades"][0]["quantity"], 10)
        self.assertTrue(queued["trades"][0]["queue"]["queue_data_available"])
        queue_rows[1]["ask_queue_ahead"] = 60
        blocked = simulate_long_only(queue_rows, [True, False], [False, False], queue_model)
        self.assertEqual(blocked["rejected_orders"][0]["reason"], "QUEUE_NOT_REACHED")
        queue_rows[1]["vi_active"] = True
        vi = simulate_long_only(queue_rows, [True, False], [False, False], queue_model)
        self.assertEqual(vi["rejected_orders"][0]["reason"], "VI_ACTIVE")

    def test_lifecycle_requires_complete_evidence_and_explicit_human_confirmation(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory))
            record = forge.create_experiment(
                StrategyDefinition("manual-gate", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8, "label_horizon_rows": 5}),
                {"dataset_id": "fixture-v1", "content_hash": "sha256:dataset"},
                {"execution_mode": "REALISTIC"},
            )
            record.status = ExperimentStatus.VALIDATION
            record.result = {
                "trade_count": 3, "dataset_hash": "sha256:dataset", "max_drawdown_pct": -10,
                "multi_timeframe": {"future_information_violations": 0},
                "benchmark_attribution": {"evidence_hash": "sha256:attribution", "geometric_excess_return_pct": 10, "information_ratio": 1.2},
                "regime_performance": {"summary": {name: {"strategy_compounded_return_pct": -5} for name in ("BULL", "BEAR", "SIDEWAYS")}},
            }
            record.validation = {"passed": True}
            forge.registry.save(record)
            blocked = forge.lifecycle_readiness(record.id)
            self.assertFalse(blocked["eligible_for_manual_paper_nomination"])
            self.assertIn("strict_walk_forward_passed", blocked["blockers"])

            record.validation.update({
                "strict_walk_forward": {"summary": {"passed": True, "temporal_leakage_detected": False}, "parameter_selection_uses_oos": False},
                "parameter_robustness": {"summary": {"robust": True}},
            })
            record.result["benchmark_attribution"].update({"geometric_excess_return_pct": -10, "information_ratio": -0.5})
            record.result["max_drawdown_pct"] = -60
            record.result["regime_performance"]["summary"]["BEAR"]["strategy_compounded_return_pct"] = -90
            forge.registry.save(record)
            forge.export(record.id)
            relative_blocked = forge.lifecycle_readiness(record.id)
            self.assertFalse(relative_blocked["eligible_for_manual_paper_nomination"])
            self.assertIn("label_horizon_purge_covered", relative_blocked["blockers"])
            self.assertIn("positive_geometric_excess_return", relative_blocked["blockers"])
            self.assertIn("positive_information_ratio", relative_blocked["blockers"])
            self.assertIn("maximum_drawdown_within_50pct", relative_blocked["blockers"])
            self.assertIn("no_catastrophic_regime_loss", relative_blocked["blockers"])

            record.result["benchmark_attribution"].update({"geometric_excess_return_pct": 10, "information_ratio": 1.2})
            record.result["max_drawdown_pct"] = -10
            record.result["regime_performance"]["summary"]["BEAR"]["strategy_compounded_return_pct"] = -5
            record.validation["strict_walk_forward"]["purge_rows"] = 5
            forge.registry.save(record)
            forge.export(record.id)
            self.assertTrue(forge.lifecycle_readiness(record.id)["eligible_for_manual_paper_nomination"])
            record.strategy.rules["label_horizon_rows"] = "tampered"
            forge.registry.save(record)
            tampered_horizon = forge.lifecycle_readiness(record.id)
            self.assertFalse(tampered_horizon["eligible_for_manual_paper_nomination"])
            self.assertIn("label_horizon_purge_covered", tampered_horizon["blockers"])
            record.strategy.rules["label_horizon_rows"] = 5
            forge.registry.save(record); forge.export(record.id)
            with self.assertRaisesRegex(ValueError, "confirmation phrase"):
                forge.review_experiment(record.id, "NOMINATE_PAPER", "reviewer", "evidence accepted")

            reviewed = forge.review_experiment(record.id, "NOMINATE_PAPER", "reviewer", "evidence accepted", CONFIRMATION)
            self.assertEqual(reviewed["experiment"]["status"], "PAPER_CANDIDATE")
            self.assertFalse(reviewed["decision"]["live_order_allowed"])
            self.assertFalse(reviewed["decision"]["readiness"]["automatic_promotion"])
            self.assertTrue(Path(reviewed["research_card"]["report"]["json_path"]).is_file())
            history = forge.lifecycle().history(record.id)
            self.assertTrue(history["chain_verified"])
            self.assertEqual(history["decision_count"], 1)

            automated = forge.create_experiment(
                StrategyDefinition("automatic-paper-gate", "1", {"type": "ma_cross", "symbol": "000660", "fast": 3, "slow": 8, "label_horizon_rows": 5}),
                {"dataset_id": "fixture-v2", "content_hash": "sha256:dataset-v2"},
                {"execution_mode": "REALISTIC"},
            )
            automated.status = ExperimentStatus.VALIDATION
            automated.result = json.loads(json.dumps(record.result))
            automated.result["dataset_hash"] = "sha256:dataset-v2"
            automated.validation = json.loads(json.dumps(record.validation))
            forge.registry.save(automated)
            forge.export(automated.id)

            auto_reviewed = forge.auto_nominate_verified_paper(
                automated.id,
                scheduler_id="verified-paper-scheduler",
                rationale="All strict evidence gates passed for Paper-only observation.",
            )

            self.assertEqual("PAPER_CANDIDATE", auto_reviewed["experiment"]["status"])
            self.assertEqual("AUTO_NOMINATE_VERIFIED_PAPER", auto_reviewed["decision"]["action"])
            self.assertEqual("automatic_verified_paper_only", auto_reviewed["decision"]["nomination_mode"])
            self.assertFalse(auto_reviewed["decision"]["manual_confirmation"])
            self.assertFalse(auto_reviewed["decision"]["automatic_live_promotion"])
            self.assertFalse(auto_reviewed["live_order_allowed"])
            auto_history = forge.lifecycle().history(automated.id)
            self.assertTrue(auto_history["chain_verified"])
            self.assertEqual(1, auto_history["decision_count"])

    def test_lifecycle_detects_audit_log_tampering(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            forge = ResearchForge.local(Path(directory))
            record = forge.create_experiment(
                StrategyDefinition("reject", "1", {"type": "ma_cross", "symbol": "005930", "fast": 3, "slow": 8}),
                {"dataset_id": "fixture-v1"}, {},
            )
            forge.review_experiment(record.id, "REJECT", "reviewer", "insufficient evidence")
            log = forge.lifecycle().log_path
            log.write_text(log.read_text(encoding="utf-8").replace("insufficient evidence", "altered evidence"), encoding="utf-8")
            self.assertFalse(forge.lifecycle().history()["chain_verified"])


if __name__ == "__main__":
    unittest.main()
