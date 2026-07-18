from __future__ import annotations

import json
import hashlib
import csv
import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .models import ExperimentRecord


def replay_payload(record: ExperimentRecord, limit: int = 500) -> dict[str, Any]:
    trades = record.result.get("trades") if isinstance(record.result.get("trades"), list) else []
    equity = record.result.get("equity_curve") if isinstance(record.result.get("equity_curve"), list) else []
    return {
        "experiment_id": record.id,
        "strategy": record.strategy.name,
        "dataset_id": record.data_snapshot.get("dataset_id"),
        "signal_timing": record.result.get("signal_timing"),
        "fill_timing": record.result.get("fill_timing"),
        "trades": trades[-max(1, min(5000, int(limit))) :],
        "equity_curve": equity[-max(1, min(5000, int(limit))) :],
        "open_positions": record.result.get("open_positions", record.result.get("open_position_quantity")),
        "read_only": True,
    }


def export_report(record: ExperimentRecord, root: Path) -> dict[str, Any]:
    root.mkdir(parents=True, exist_ok=True)
    payload = record.to_dict()
    json_path = root / f"{record.id}.json"
    markdown_path = root / f"{record.id}.md"
    excel_path = root / f"{record.id}.xlsx"
    csv_path = root / f"{record.id}.csv"
    manifest_path = root / f"{record.id}.manifest.json"
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    evidence = {
        "schema_version": 1,
        "experiment_id": record.id,
        "experiment_hash": f"sha256:{hashlib.sha256(canonical.encode()).hexdigest()}",
        "dataset_id": record.data_snapshot.get("dataset_id"),
        "dataset_hash": record.data_snapshot.get("content_hash") or record.data_snapshot.get("cache_hash"),
        "code_version": record.code_version,
        "adapter": record.backtest_adapter,
        "random_seed": record.random_seed,
        "validation_passed": bool(record.validation.get("passed")),
        "recommendation": record.validation.get("recommendation", "RESEARCH_ONLY"),
        "research_only": True,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    _atomic_write(json_path, json.dumps({"experiment": payload, "evidence": evidence}, ensure_ascii=False, indent=2, sort_keys=True))
    _atomic_write(markdown_path, _markdown(record))
    _excel(record, excel_path, evidence)
    _atomic_write(csv_path, _csv(record))
    evidence["artifacts"] = {
        path.name: {"sha256": _hash(path), "bytes": path.stat().st_size}
        for path in (json_path, markdown_path, excel_path, csv_path)
    }
    _atomic_write(manifest_path, json.dumps(evidence, ensure_ascii=False, indent=2, sort_keys=True))
    return {
        "experiment_id": record.id,
        "json_path": str(json_path),
        "markdown_path": str(markdown_path),
        "excel_path": str(excel_path),
        "csv_path": str(csv_path),
        "manifest_path": str(manifest_path),
        "experiment_hash": evidence["experiment_hash"],
        "artifact_hashes": evidence["artifacts"],
        "formats": ["json", "markdown", "xlsx", "csv"],
    }


def verify_report(record: ExperimentRecord, root: Path) -> dict[str, Any]:
    manifest_path = root / f"{record.id}.manifest.json"
    if not manifest_path.is_file():
        return {"ok": False, "experiment_id": record.id, "errors": ["manifest_missing"]}
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    canonical = json.dumps(record.to_dict(), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    expected_experiment = f"sha256:{hashlib.sha256(canonical.encode()).hexdigest()}"
    errors = []
    if manifest.get("experiment_hash") != expected_experiment:
        errors.append("experiment_hash_mismatch")
    for name, evidence in (manifest.get("artifacts") or {}).items():
        path = root / Path(str(name)).name
        if not path.is_file():
            errors.append(f"artifact_missing:{name}")
        elif _hash(path) != str((evidence or {}).get("sha256")):
            errors.append(f"artifact_hash_mismatch:{name}")
    if not manifest.get("artifacts"):
        errors.append("artifact_evidence_missing")
    return {"ok": not errors, "experiment_id": record.id, "errors": errors, "checked_artifacts": len(manifest.get("artifacts") or {}), "experiment_hash": expected_experiment}


def _markdown(record: ExperimentRecord) -> str:
    result, validation = record.result, record.validation
    blockers = validation.get("blockers") if isinstance(validation.get("blockers"), list) else []
    walk = validation.get("walk_forward") if isinstance(validation.get("walk_forward"), dict) else {}
    strict_walk = validation.get("strict_walk_forward") if isinstance(validation.get("strict_walk_forward"), dict) else {}
    robust = validation.get("parameter_robustness") if isinstance(validation.get("parameter_robustness"), dict) else {}
    regimes = result.get("regime_performance") if isinstance(result.get("regime_performance"), dict) else {}
    attribution = result.get("benchmark_attribution") if isinstance(result.get("benchmark_attribution"), dict) else {}
    lines = [
        f"# Research Forge Experiment {record.id}",
        "",
        f"- Strategy: {record.strategy.name} ({record.strategy.version})",
        f"- Status: {record.status.value}",
        f"- Dataset: {record.data_snapshot.get('dataset_id', '-')}",
        f"- Adapter: {record.backtest_adapter}",
        f"- Total return: {result.get('total_return_pct', '-')}%",
        f"- Maximum drawdown: {result.get('max_drawdown_pct', '-')}%",
        f"- Trades: {result.get('trade_count', 0)}",
        f"- Validation: {'PASS' if validation.get('passed') else 'BLOCKED'}",
        f"- Recommendation: {validation.get('recommendation', 'RESEARCH_ONLY')}",
        f"- Blockers: {', '.join(str(value) for value in blockers) if blockers else '-'}",
        "",
        "## Robustness",
        "",
        f"- Walk-forward: {json.dumps(walk.get('summary', {}), ensure_ascii=False)}",
        f"- Strict walk-forward: {json.dumps(strict_walk.get('summary', {}), ensure_ascii=False)}",
        f"- Strict boundary exclusions: purge_days={strict_walk.get('purge_days', 0)}, embargo_days={strict_walk.get('embargo_days', 0)}, purge_rows={strict_walk.get('purge_rows', 0)}, embargo_rows={strict_walk.get('embargo_rows', 0)}, label_horizon_rows={strict_walk.get('declared_label_horizon_rows', 0)}",
        f"- Parameter robustness: {json.dumps(robust.get('summary', {}), ensure_ascii=False)}",
        "",
        "## Market regimes",
        "",
        f"- Method: {regimes.get('method', '-')}",
        f"- Lookback bars: {regimes.get('lookback_bars', '-')}",
        f"- Threshold: {regimes.get('threshold_pct', '-')}%",
        f"- Results: {json.dumps(regimes.get('summary', {}), ensure_ascii=False)}",
        "",
        "## Benchmark attribution",
        "",
        f"- Benchmark: {attribution.get('benchmark', '-')}",
        f"- Benchmark return: {attribution.get('benchmark_return_pct', '-')}%",
        f"- Geometric excess return: {attribution.get('geometric_excess_return_pct', '-')}%",
        f"- Beta: {attribution.get('beta', '-')}",
        f"- Correlation: {attribution.get('correlation', '-')}",
        f"- Tracking error: {attribution.get('annualized_tracking_error_pct', '-')}%",
        f"- Information ratio: {attribution.get('information_ratio', '-')}",
        "",
        "## Evidence",
        "",
        f"- Code version: {record.code_version}",
        f"- Random seed: {record.random_seed}",
        f"- Dataset hash: {record.data_snapshot.get('content_hash') or record.data_snapshot.get('cache_hash') or '-'}",
        f"- Execution model: `{json.dumps(record.execution_model, ensure_ascii=False, sort_keys=True)}`",
        "",
        "This report is research-only and cannot authorize or submit a live order.",
        "",
    ]
    return "\n".join(lines)


def _atomic_write(path: Path, text: str) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(text, encoding="utf-8")
    temporary.replace(path)


def _excel(record: ExperimentRecord, path: Path, evidence: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("Excel reports require openpyxl: pip install .[reports]") from exc
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    rows = [
        ("Experiment ID", record.id), ("Strategy", record.strategy.name),
        ("Status", record.status.value), ("Dataset", record.data_snapshot.get("dataset_id")),
        ("Total return (%)", record.result.get("total_return_pct")),
        ("Max drawdown (%)", record.result.get("max_drawdown_pct")),
        ("Trade count", record.result.get("trade_count", 0)),
        ("Validation", "PASS" if record.validation.get("passed") else "BLOCKED"),
        ("Recommendation", record.validation.get("recommendation", "RESEARCH_ONLY")),
        ("Research only", True),
    ]
    summary.append(["Field", "Value"])
    for row in rows: summary.append([_cell(value) for value in row])
    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_sheet.append(["Field", "Value"])
    for key, value in evidence.items(): evidence_sheet.append([key, _cell(value)])
    validation_sheet = workbook.create_sheet("Validation")
    validation_sheet.append(["Field", "Value"])
    for key, value in sorted(record.validation.items()): validation_sheet.append([key, _cell(value)])
    strict_sheet = workbook.create_sheet("Strict Walk Forward")
    strict = record.validation.get("strict_walk_forward") if isinstance(record.validation.get("strict_walk_forward"), dict) else {}
    strict_sheet.append(["Fold", "Train start", "Train end", "Test start", "Test end", "Purge days", "Embargo days", "Purge rows", "Embargo rows", "Train rows", "OOS rows", "OOS return (%)", "OOS MDD (%)", "OOS trades", "Leakage"])
    for window in strict.get("windows", []) if isinstance(strict.get("windows"), list) else []:
        strict_sheet.append([window.get("fold"), window.get("train_start"), window.get("train_end"), window.get("test_start"), window.get("test_end"), window.get("purge_days", 0), window.get("embargo_days", 0), window.get("purge_rows", 0), window.get("embargo_rows", 0), (window.get("selected") or {}).get("train_row_count"), window.get("oos_row_count"), window.get("oos_return_pct"), window.get("oos_mdd_pct"), window.get("oos_trade_count"), window.get("temporal_leakage")])
    strict_sheet.append([])
    strict_sheet.append(["Setting", "Value"])
    for key in ("requested_purge_rows", "declared_label_horizon_rows", "purge_rows", "embargo_rows", "purge_days", "embargo_days"):
        strict_sheet.append([key, strict.get(key, 0)])
    regimes_sheet = workbook.create_sheet("Regimes")
    regimes_sheet.append(["Regime", "Bar count", "Compounded return (%)", "Average bar return (%)", "Best bar (%)", "Worst bar (%)"])
    regime_summary = ((record.result.get("regime_performance") or {}).get("summary") or {})
    for name in ("BULL", "BEAR", "SIDEWAYS"):
        value = regime_summary.get(name) or {}; regimes_sheet.append([name, value.get("bar_count", 0), value.get("strategy_compounded_return_pct", 0), value.get("strategy_average_bar_return_pct", 0), value.get("best_bar_return_pct", 0), value.get("worst_bar_return_pct", 0)])
    attribution_sheet = workbook.create_sheet("Attribution")
    attribution_sheet.append(["Field", "Value"])
    for key, value in sorted((record.result.get("benchmark_attribution") or {}).items()): attribution_sheet.append([key, _cell(value)])
    trades_sheet = workbook.create_sheet("Trades")
    trades = record.result.get("trades") if isinstance(record.result.get("trades"), list) else []
    columns = sorted({key for trade in trades if isinstance(trade, dict) for key in trade})
    trades_sheet.append(columns or ["No trades"])
    for trade in trades: trades_sheet.append([_cell(trade.get(key)) for key in columns])
    for sheet in workbook.worksheets:
        for cell in sheet[1]: cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    temporary = path.with_suffix(".xlsx.tmp")
    workbook.save(temporary)
    temporary.replace(path)


def _csv(record: ExperimentRecord) -> str:
    stream = io.StringIO(newline=""); writer = csv.writer(stream, lineterminator="\n")
    writer.writerow(["category", "key", "value"])
    summary = {"experiment_id": record.id, "strategy": record.strategy.name, "status": record.status.value, "dataset_id": record.data_snapshot.get("dataset_id"), "total_return_pct": record.result.get("total_return_pct"), "max_drawdown_pct": record.result.get("max_drawdown_pct"), "trade_count": record.result.get("trade_count", 0), "validation_passed": bool(record.validation.get("passed")), "recommendation": record.validation.get("recommendation", "RESEARCH_ONLY"), "research_only": True}
    for key, value in summary.items(): writer.writerow(["summary", _csv_cell(key), _csv_cell(value)])
    strict_walk = record.validation.get("strict_walk_forward") if isinstance(record.validation.get("strict_walk_forward"), dict) else {}
    for key in ("purge_days", "embargo_days", "purge_rows", "embargo_rows", "requested_purge_rows", "declared_label_horizon_rows"):
        writer.writerow(["strict_walk_forward", key, _csv_cell(strict_walk.get(key, 0))])
    writer.writerow(["strict_walk_forward", "summary", _csv_cell(strict_walk.get("summary", {}))])
    for name, value in sorted((((record.result.get("regime_performance") or {}).get("summary") or {}).items())):
        writer.writerow(["regime", _csv_cell(name), _csv_cell(value)])
    for key, value in sorted((record.result.get("benchmark_attribution") or {}).items()):
        writer.writerow(["benchmark_attribution", _csv_cell(key), _csv_cell(value)])
    for index, trade in enumerate(record.result.get("trades") if isinstance(record.result.get("trades"), list) else []):
        writer.writerow(["trade", str(index + 1), _csv_cell(trade)])
    return stream.getvalue()


def _csv_cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)): value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")): return "'" + value
    return value


def _cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
